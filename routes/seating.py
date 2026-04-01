"""
Seating Plan – upload file tavoli, check vs DB, CRUD assegnazioni.
"""
import io
import unicodedata
from datetime import datetime

import pandas as pd
from flask import (Blueprint, render_template, request, jsonify, session,
                   flash, redirect, url_for, send_file)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models import db
from models.models import RoomingList, TableAssignment

seating_bp = Blueprint('seating', __name__)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _norm(s):
    """Normalizza stringa per confronto: lowercase, strip, rimuovi accenti."""
    if not s:
        return ''
    s = str(s).strip().lower()
    # Decomponi caratteri accentati e rimuovi combining marks
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return s


def _match_participants(rows, batch_id):
    """
    Per ogni riga del file, cerca il miglior match nella rooming_list.
    Restituisce lista di dict con: file_data, match_type, candidate(s).
    """
    # Carica tutti i partecipanti del batch più recente
    if batch_id:
        participants = RoomingList.query.filter_by(import_batch=batch_id).all()
    else:
        # Prendi l'ultimo batch
        latest = (db.session.query(RoomingList.import_batch)
                  .order_by(RoomingList.imported_at.desc()).first())
        if not latest:
            return [], None
        batch_id = latest[0]
        participants = RoomingList.query.filter_by(import_batch=batch_id).all()

    # Indici per lookup veloce
    by_name = {}       # (norm_last, norm_first) -> [p, ...]
    by_email = {}      # norm_email -> p
    by_last = {}       # norm_last -> [p, ...]
    by_company = {}    # (norm_last, norm_company) -> [p, ...]

    for p in participants:
        nl = _norm(p.last_name)
        nf = _norm(p.first_name)
        ne = _norm(p.email)
        nc = _norm(p.company_name)

        key = (nl, nf)
        by_name.setdefault(key, []).append(p)
        if ne:
            by_email[ne] = p
        if nl:
            by_last.setdefault(nl, []).append(p)
        if nl and nc:
            by_company.setdefault((nl, nc), []).append(p)

    results = []
    for row in rows:
        file_last = _norm(row.get('last_name', ''))
        file_first = _norm(row.get('first_name', ''))
        file_email = _norm(row.get('email', ''))
        file_company = _norm(row.get('company', ''))

        match_type = 'not_found'
        candidates = []

        # 1) Match esatto nome+cognome
        exact = by_name.get((file_last, file_first))
        if exact:
            match_type = 'exact'
            candidates = exact
        # 2) Match per email
        elif file_email and file_email in by_email:
            match_type = 'email'
            candidates = [by_email[file_email]]
        # 3) Match cognome + azienda
        elif file_last and file_company:
            comp_match = by_company.get((file_last, file_company))
            if comp_match:
                match_type = 'partial'
                candidates = comp_match
        # 4) Match solo cognome (possibili duplicati)
        if not candidates and file_last:
            last_match = by_last.get(file_last, [])
            if len(last_match) == 1:
                match_type = 'partial'
                candidates = last_match
            elif len(last_match) > 1:
                match_type = 'ambiguous'
                candidates = last_match

        results.append({
            'file_data': row,
            'match_type': match_type,
            'candidates': [
                {
                    'id': c.id,
                    'internal_reference': c.internal_reference,
                    'name': f"{c.title or ''} {c.first_name or ''} {c.last_name or ''}".strip(),
                    'company': c.company_name or '',
                    'hotel': c.hotel or '',
                    'email': c.email or '',
                }
                for c in candidates
            ],
        })

    return results, batch_id


def _parse_upload(file_storage):
    """Parsa il file Excel caricato e restituisce lista di dict."""
    df = pd.read_excel(file_storage, engine='openpyxl')

    # Normalizza nomi colonne
    col_map = {}
    for c in df.columns:
        cl = str(c).strip().upper()
        if 'LAST' in cl and 'NAME' in cl:
            col_map[c] = 'last_name'
        elif 'FIRST' in cl and 'NAME' in cl:
            col_map[c] = 'first_name'
        elif 'TABLE' in cl:
            col_map[c] = 'table_number'
        elif cl == 'TITLE':
            col_map[c] = 'title'
        elif cl == 'POSITION':
            col_map[c] = 'position'
        elif cl == 'EMAIL':
            col_map[c] = 'email'
        elif cl == 'PHONE':
            col_map[c] = 'phone'
        elif cl == 'COMPANY':
            col_map[c] = 'company'
        elif cl == 'CATEGORY':
            col_map[c] = 'category'

    df = df.rename(columns=col_map)

    rows = []
    for _, r in df.iterrows():
        row = {}
        for field in ['last_name', 'first_name', 'table_number', 'title',
                       'position', 'email', 'phone', 'company', 'category']:
            val = r.get(field)
            if pd.isna(val):
                row[field] = ''
            else:
                row[field] = str(val).strip()
        # table_number: rimuovi .0 se intero
        tn = row['table_number']
        if tn.endswith('.0'):
            row['table_number'] = tn[:-2]
        rows.append(row)

    return rows


# ── Pagina principale ────────────────────────────────────────────────────────

@seating_bp.route('/seating')
def seating_page():
    # Statistiche correnti
    total = TableAssignment.query.count()
    tables = (db.session.query(TableAssignment.table_number,
                               db.func.count(TableAssignment.id))
              .group_by(TableAssignment.table_number)
              .order_by(TableAssignment.table_number)
              .all())
    return render_template('seating.html', total=total, tables=tables)


# ── Upload & Preview ─────────────────────────────────────────────────────────

@seating_bp.route('/seating/upload', methods=['POST'])
def seating_upload():
    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        flash('Carica un file Excel (.xlsx).', 'error')
        return redirect(url_for('seating.seating_page'))

    rows = _parse_upload(f)

    # Trova ultimo batch rooming
    latest = (db.session.query(RoomingList.import_batch)
              .order_by(RoomingList.imported_at.desc()).first())
    batch_id = latest[0] if latest else None

    results, batch_id = _match_participants(rows, batch_id)

    # Statistiche
    stats = {
        'total': len(results),
        'exact': sum(1 for r in results if r['match_type'] == 'exact'),
        'partial': sum(1 for r in results if r['match_type'] == 'partial'),
        'email': sum(1 for r in results if r['match_type'] == 'email'),
        'ambiguous': sum(1 for r in results if r['match_type'] == 'ambiguous'),
        'not_found': sum(1 for r in results if r['match_type'] == 'not_found'),
    }

    return render_template('seating.html',
                           preview=True,
                           results=results,
                           stats=stats,
                           batch_id=batch_id,
                           total=TableAssignment.query.count(),
                           tables=[])


# ── Conferma import ──────────────────────────────────────────────────────────

@seating_bp.route('/seating/confirm-import', methods=['POST'])
def seating_confirm_import():
    data = request.get_json()
    if not data or 'assignments' not in data:
        return jsonify(ok=False, error='Dati mancanti'), 400

    batch_label = f"seating_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    username = session.get('username', 'system')
    count = 0

    for item in data['assignments']:
        ta = TableAssignment(
            rooming_list_id=item.get('rooming_list_id') or None,
            table_number=str(item['table_number']),
            last_name=item.get('last_name', ''),
            first_name=item.get('first_name', ''),
            company=item.get('company', ''),
            category=item.get('category', ''),
            title=item.get('title', ''),
            position=item.get('position', ''),
            email=item.get('email', ''),
            phone=item.get('phone', ''),
            import_batch=batch_label,
            modified_by=username,
        )
        db.session.add(ta)
        count += 1

    db.session.commit()
    return jsonify(ok=True, count=count, batch=batch_label)


# ── API: lista tavoli con partecipanti ───────────────────────────────────────

@seating_bp.route('/api/seating/tables')
def api_seating_tables():
    """Restituisce tutti i tavoli con partecipanti, ordinabili per tavolo o hotel."""
    sort_by = request.args.get('sort', 'table')  # table | hotel | name

    query = TableAssignment.query

    # Join opzionale con rooming_list per avere hotel
    assignments = (db.session.query(TableAssignment, RoomingList)
                   .outerjoin(RoomingList, TableAssignment.rooming_list_id == RoomingList.id)
                   .all())

    tables_dict = {}
    for ta, rl in assignments:
        tn = ta.table_number
        tables_dict.setdefault(tn, []).append({
            'id': ta.id,
            'table_number': tn,
            'last_name': ta.last_name or '',
            'first_name': ta.first_name or '',
            'company': ta.company or '',
            'category': ta.category or '',
            'title': ta.title or '',
            'email': ta.email or '',
            'phone': ta.phone or '',
            'position': ta.position or '',
            'hotel': rl.hotel if rl else '',
            'rooming_list_id': ta.rooming_list_id,
            'internal_reference': rl.internal_reference if rl else '',
            'registration_state': rl.registration_state if rl else '',
        })

    # Ordina tavoli numericamente
    def table_sort_key(tn):
        try:
            return (0, int(tn))
        except ValueError:
            return (1, tn)

    result = []
    for tn in sorted(tables_dict.keys(), key=table_sort_key):
        people = tables_dict[tn]
        if sort_by == 'hotel':
            people.sort(key=lambda x: (x['hotel'] or 'zzz', x['last_name']))
        else:
            people.sort(key=lambda x: x['last_name'])
        result.append({'table_number': tn, 'count': len(people), 'participants': people})

    return jsonify(tables=result, total=sum(len(t['participants']) for t in result))


# ── API: aggiorna assegnazione (cambio tavolo) ──────────────────────────────

@seating_bp.route('/api/seating/update', methods=['POST'])
def api_seating_update():
    data = request.get_json()
    ta_id = data.get('id')
    if not ta_id:
        return jsonify(ok=False, error='ID mancante'), 400

    ta = TableAssignment.query.get(ta_id)
    if not ta:
        return jsonify(ok=False, error='Assegnazione non trovata'), 404

    if 'table_number' in data:
        ta.table_number = str(data['table_number'])
    if 'last_name' in data:
        ta.last_name = data['last_name']
    if 'first_name' in data:
        ta.first_name = data['first_name']
    if 'company' in data:
        ta.company = data['company']
    if 'category' in data:
        ta.category = data['category']

    ta.modified_by = session.get('username', 'system')
    ta.modified_at = datetime.utcnow()

    # ── Sync → RoomingList (se collegato) ────────────────────────────────
    SYNC_FIELDS = {'last_name': 'last_name', 'first_name': 'first_name',
                   'company': 'company_name'}
    synced_fields = [f for f in SYNC_FIELDS if f in data]
    if ta.rooming_list_id and synced_fields:
        rl = RoomingList.query.get(ta.rooming_list_id)
        if rl:
            for ta_field in synced_fields:
                setattr(rl, SYNC_FIELDS[ta_field], data[ta_field])

    db.session.commit()

    return jsonify(ok=True)


# ── API: aggiungi assegnazione manuale ───────────────────────────────────────

@seating_bp.route('/api/seating/add', methods=['POST'])
def api_seating_add():
    data = request.get_json()
    if not data or not data.get('table_number'):
        return jsonify(ok=False, error='Numero tavolo obbligatorio'), 400

    ta = TableAssignment(
        rooming_list_id=data.get('rooming_list_id') or None,
        table_number=str(data['table_number']),
        last_name=data.get('last_name', ''),
        first_name=data.get('first_name', ''),
        company=data.get('company', ''),
        category=data.get('category', ''),
        title=data.get('title', ''),
        position=data.get('position', ''),
        email=data.get('email', ''),
        phone=data.get('phone', ''),
        modified_by=session.get('username', 'system'),
    )
    db.session.add(ta)
    db.session.commit()

    return jsonify(ok=True, id=ta.id)


# ── API: elimina assegnazione ────────────────────────────────────────────────

@seating_bp.route('/api/seating/delete', methods=['POST'])
def api_seating_delete():
    data = request.get_json()
    ta_id = data.get('id')
    if not ta_id:
        return jsonify(ok=False, error='ID mancante'), 400

    ta = TableAssignment.query.get(ta_id)
    if not ta:
        return jsonify(ok=False, error='Non trovato'), 404

    db.session.delete(ta)
    db.session.commit()
    return jsonify(ok=True)


# ── API: elimina TUTTI i tavoli ──────────────────────────────────────────────

@seating_bp.route('/api/seating/delete-all', methods=['POST'])
def api_seating_delete_all():
    count = TableAssignment.query.delete()
    db.session.commit()
    return jsonify(ok=True, deleted=count)


# ── API: cerca partecipanti rooming per aggiunta manuale ─────────────────────

@seating_bp.route('/api/seating/search-rooming', methods=['POST'])
def api_seating_search_rooming():
    data = request.get_json()
    q = (data.get('q') or '').strip()
    if len(q) < 2:
        return jsonify(results=[])

    latest = (db.session.query(RoomingList.import_batch)
              .order_by(RoomingList.imported_at.desc()).first())
    if not latest:
        return jsonify(results=[])

    pattern = f'%{q}%'
    matches = (RoomingList.query
               .filter(RoomingList.import_batch == latest[0])
               .filter(db.or_(
                   RoomingList.last_name.ilike(pattern),
                   RoomingList.first_name.ilike(pattern),
                   RoomingList.company_name.ilike(pattern),
                   RoomingList.email.ilike(pattern),
               ))
               .limit(20)
               .all())

    return jsonify(results=[
        {
            'id': p.id,
            'name': f"{p.title or ''} {p.first_name or ''} {p.last_name or ''}".strip(),
            'company': p.company_name or '',
            'hotel': p.hotel or '',
            'email': p.email or '',
            'internal_reference': p.internal_reference or '',
        }
        for p in matches
    ])


# ── Export Excel ─────────────────────────────────────────────────────────────

@seating_bp.route('/seating/export')
def seating_export():
    sort_by = request.args.get('sort', 'table')

    assignments = (db.session.query(TableAssignment, RoomingList)
                   .outerjoin(RoomingList, TableAssignment.rooming_list_id == RoomingList.id)
                   .all())

    rows = []
    for ta, rl in assignments:
        rows.append({
            'Table': ta.table_number,
            'Title': ta.title or '',
            'Last Name': ta.last_name or '',
            'First Name': ta.first_name or '',
            'Company': ta.company or '',
            'Category': ta.category or '',
            'Hotel': rl.hotel if rl else '',
            'Check': 'CXL' if rl and (rl.registration_state or '').strip().upper() == 'CXL'
                        else ('No Rooming' if not rl else ''),
            'Email': ta.email or '',
            'Phone': ta.phone or '',
            'Position': ta.position or '',
        })

    def sort_key(r):
        try:
            tn = int(r['Table'])
        except (ValueError, TypeError):
            tn = 9999
        if sort_by == 'hotel':
            return (r.get('Hotel') or 'zzz', tn, r['Last Name'])
        return (tn, r['Last Name'])

    rows.sort(key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Seating Plan'

    headers = ['Table', 'Title', 'Last Name', 'First Name', 'Company',
               'Category', 'Hotel', 'Check', 'Email', 'Phone', 'Position']

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(h, ''))
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26
                             else 'A'].width = 18

    # ── Foglio 2: partecipanti rooming NON assegnati a un tavolo ─────────
    assigned_ids = {ta.rooming_list_id for ta, rl in assignments if ta.rooming_list_id}

    latest = (db.session.query(RoomingList.import_batch)
              .order_by(RoomingList.imported_at.desc()).first())
    if latest:
        all_rooming = RoomingList.query.filter_by(import_batch=latest[0]).all()
        not_assigned = [p for p in all_rooming
                        if p.id not in assigned_ids
                        and (p.registration_state or '').strip().upper() != 'CXL']
        not_assigned.sort(key=lambda p: (p.last_name or '', p.first_name or ''))

        if not_assigned:
            ws2 = wb.create_sheet('Non Importati')
            h2 = ['Title', 'Last Name', 'First Name', 'Company', 'Hotel',
                   'Check', 'Email', 'Phone']
            warn_fill = PatternFill(start_color='B45309', end_color='B45309', fill_type='solid')

            for col_idx, h in enumerate(h2, 1):
                cell = ws2.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = warn_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            for row_idx, p in enumerate(not_assigned, 2):
                vals = [p.title or '', p.last_name or '', p.first_name or '',
                        p.company_name or '', p.hotel or '',
                        'No Table', p.email or '', p.phone or '']
                for col_idx, v in enumerate(vals, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=v)
                    cell.border = thin_border

            for col_idx in range(1, len(h2) + 1):
                ws2.column_dimensions[chr(64 + col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"seating_plan_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
