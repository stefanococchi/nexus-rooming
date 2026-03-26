import os
import io
import re
from datetime import datetime, date

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, send_file, jsonify)

from models import db
from models.models import RoomingList

rooming_bp = Blueprint('rooming', __name__)

# ── Mappa colonne XLS ────────────────────────────────────────────────────────

COL_MAP = {
    0:  'registration_state',
    1:  'latest_changes',
    2:  'hotel',
    3:  'upgrade',
    4:  'participant_display',
    5:  'billing',
    6:  'company_name',
    7:  'company_country',
    8:  'nexus_bd',
    9:  'is_parent_manager',
    10: 'registered_colleagues',
    11: 'internal_reference',
    12: 'internal_parent_reference',
    13: 'ean8_barcode',
    14: 'participant_number',
    15: 'external_reference',
    16: 'delegation_key',
    17: 'status_vp_bd',
    18: 'status_organisator',
    19: 'status_board_nai',
    20: 'status_climate_day',
    21: 'status_prospective_council',
    22: 'status_spouse',
    23: 'comment',
    24: 'title',
    25: 'last_name',
    26: 'first_name',
    27: 'job_position',
    28: 'email',
    29: 'phone',
    30: 'prospective_title',
    31: 'prospective_response',
    32: 'night_no_need',
    33: 'night_sat_28mar',
    34: 'night_sun_29mar',
    35: 'night_mon_30mar',
    36: 'night_tue_31mar',
    37: 'night_wed_1apr',
    38: 'night_thu_2apr',
    39: 'night_fri_3apr',
    40: 'night_sat_4apr',
    41: 'night_other',
    42: 'diet_restrictions',
    43: 'arrival_mode',
    44: 'need_smooth_checkin',
    45: 'need_visa',
    46: 'visa_birth_date',
    47: 'visa_birth_place',
    48: 'visa_passport',
    49: 'visa_delivery_date',
    50: 'visa_expiration_date',
    51: 'visa_company_address',
    52: 'company_category',
    53: 'company_subcategory',
}

DATE_FIELDS = {'visa_birth_date', 'visa_delivery_date', 'visa_expiration_date',
               'change_date', 'file_timestamp', 'check_in', 'check_out'}
INT_FIELDS  = {'registered_colleagues', 'participant_number'}

# Colonne per l'Excel di output hotel (sequenza standard)
EXPORT_COLS = [
    ('title',                'Title'),
    ('last_name',            'Last Name'),
    ('first_name',           'First Name'),
    ('comment',              'Comment'),
    ('email',                'Email'),
    ('phone',                'Phone'),
    ('billing',              'Billing'),
    ('company_category',     'Category'),
    ('upgrade',              'Upgrade'),
    ('arrival_mode',         'Arrival Mode'),
    ('check_in',             'Check In'),
    ('check_out',            'Check Out'),
    ('need_visa',            'Visa?'),
    ('visa_birth_date',      'Birth Date'),
    ('visa_birth_place',     'Birth Place'),
    ('visa_passport',        'Passport N.'),
    ('visa_delivery_date',   'Doc Issued'),
    ('visa_expiration_date', 'Doc Expiry'),
    ('latest_changes',       'Latest Changes'),
    ('change_type',          'Change Type'),
    ('change_date',          'Change Date'),
]

VIRTUAL_FIELDS = {'check_in', 'check_out'}
DIFF_FIELDS = [
    # Campi originali da EXPORT_COLS
    'title', 'last_name', 'first_name', 'comment', 'email', 'phone',
    'billing', 'company_category', 'upgrade', 'arrival_mode',
    'need_visa', 'visa_birth_date', 'visa_birth_place', 'visa_passport',
    'visa_delivery_date', 'visa_expiration_date',
    'latest_changes', 'change_type', 'change_date',
    # Aggiunti: hotel, notti, stato
    'registration_state', 'hotel',
    'night_no_need',
    'night_sat_28mar', 'night_sun_29mar', 'night_mon_30mar', 'night_tue_31mar',
    'night_wed_1apr', 'night_thu_2apr', 'night_fri_3apr', 'night_sat_4apr',
]

COL_WIDTHS = {
    'title': 6, 'last_name': 18, 'first_name': 16,
    'comment': 35, 'email': 28, 'phone': 16,
    'billing': 14, 'upgrade': 12,
    'check_in': 12, 'check_out': 12,
    'need_visa': 7, 'visa_birth_date': 12,
    'visa_birth_place': 16, 'visa_passport': 14,
    'visa_delivery_date': 12, 'visa_expiration_date': 12,
    'latest_changes': 28, 'change_type': 16, 'change_date': 12,
    # altri usati nell'export dinamico
    'company_name': 28, 'company_country': 12, 'job_position': 30,
    'company_category': 14, 'company_subcategory': 14,
    'registration_state': 10, 'hotel': 20,
    'night_no_need': 8, 'night_sat_28mar': 8, 'night_sun_29mar': 8,
    'night_mon_30mar': 8, 'night_tue_31mar': 8, 'night_wed_1apr': 8,
    'night_thu_2apr': 8, 'night_fri_3apr': 8, 'night_sat_4apr': 8,
    'arrival_mode': 10, 'need_smooth_checkin': 9, 'diet_restrictions': 20,
    'visa_company_address': 30, 'nexus_bd': 20, 'internal_reference': 14,
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def clean_value(field, val):
    if pd.isna(val) if not isinstance(val, (list, dict)) else False:
        return None
    if field in DATE_FIELDS:
        if isinstance(val, (datetime, date)):
            return val if isinstance(val, date) else val.date()
        try:
            return pd.to_datetime(val).date()
        except Exception:
            return None
    if field in INT_FIELDS:
        try:
            return int(val)
        except Exception:
            return None
    s = str(val).strip()
    return s if s and s.lower() not in ('nan', 'none') else None


def make_batch_id(filename):
    ts    = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = re.sub(r'[^\w\-_. ]', '_', filename)
    return f"{ts}_{fname}"


def get_batches():
    rows = db.session.execute(db.text("""
        SELECT import_batch,
               COUNT(*)        AS total,
               MIN(imported_at) AS imported_at
        FROM rooming_list
        GROUP BY import_batch
        ORDER BY imported_at DESC
    """)).fetchall()
    return rows


def fetch_batch_dict(batch_id):
    rows = RoomingList.query.filter_by(import_batch=batch_id).all()
    result = {}
    for r in rows:
        key = r.internal_reference or f'__noref_{r.last_name}_{r.first_name}'
        result[key] = r
    return result


def compute_diff(current_dict, previous_dict):
    diff = {}
    for ref, row in current_dict.items():
        if ref not in previous_dict:
            diff[ref] = ['NEW']
        else:
            prev = previous_dict[ref]
            changed = [f for f in DIFF_FIELDS
                       if str(getattr(row, f) or '').strip() !=
                          str(getattr(prev, f) or '').strip()]
            if changed:
                diff[ref] = changed
    return diff


from datetime import date as _date_type

NIGHT_DATES = [
    ('night_sat_28mar', _date_type(2026, 3, 28)),
    ('night_sun_29mar', _date_type(2026, 3, 29)),
    ('night_mon_30mar', _date_type(2026, 3, 30)),
    ('night_tue_31mar', _date_type(2026, 3, 31)),
    ('night_wed_1apr',  _date_type(2026, 4,  1)),
    ('night_thu_2apr',  _date_type(2026, 4,  2)),
    ('night_fri_3apr',  _date_type(2026, 4,  3)),
    ('night_sat_4apr',  _date_type(2026, 4,  4)),
]

def get_checkin(row):
    for field, d in NIGHT_DATES:
        if getattr(row, field):
            return d
    return None

def get_checkout(row):
    from datetime import timedelta
    last = None
    for field, d in NIGHT_DATES:
        if getattr(row, field):
            last = d
    return last + timedelta(days=1) if last else None


def fmt_date(val):
    if val is None:
        return ''
    if hasattr(val, 'strftime'):
        return val.strftime('%d/%m/%Y')
    return str(val)


def cell_val(field, row):
    if field == 'check_in':
        return fmt_date(get_checkin(row))
    if field == 'check_out':
        return fmt_date(get_checkout(row))
    if field == 'spouse_name':
        return ''
    val = getattr(row, field, None)
    if val is None:
        return ''
    if field in DATE_FIELDS:
        return fmt_date(val)
    s = str(val).strip()
    return '' if s.lower() in ('none', 'nan') else s


def build_hotel_excel(hotel_name, rows, diff, batch_id, prev_batch_id):
    from models.models import HotelContract
    from datetime import date as _date_type

    wb = openpyxl.Workbook()
    # Rinomina il foglio default che verrà usato come Pivot
    ws_pivot = wb.active
    ws_pivot.title = 'Pivot'

    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', start_color='1F3864')
    tot_font  = Font(name='Calibri', bold=True, size=10)
    tot_fill  = PatternFill('solid', start_color='D9E1F2')
    norm_font_p = Font(name='Calibri', size=10)
    center_p  = Alignment(horizontal='center', vertical='center')
    left_p    = Alignment(horizontal='left',   vertical='center')
    thin_p    = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'),  bottom=Side(style='thin'))

    PIVOT_NIGHTS = [
        ('night_sat_28mar', '28-mar', _date_type(2026, 3, 28)),
        ('night_sun_29mar', '29-mar', _date_type(2026, 3, 29)),
        ('night_mon_30mar', '30-mar', _date_type(2026, 3, 30)),
        ('night_tue_31mar', '31-mar', _date_type(2026, 3, 31)),
        ('night_wed_1apr',  '1-apr',  _date_type(2026, 4,  1)),
        ('night_thu_2apr',  '2-apr',  _date_type(2026, 4,  2)),
        ('night_fri_3apr',  '3-apr',  _date_type(2026, 4,  3)),
        ('night_sat_4apr',  '4-apr',  _date_type(2026, 4,  4)),
    ]

    # Titolo pivot
    ws_pivot.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PIVOT_NIGHTS)+2)
    tc = ws_pivot.cell(row=1, column=1, value=f'CONFIRMED vs BY CONTRACT — {hotel_name.upper()}')
    tc.font = Font(name='Calibri', bold=True, size=13, color='1F3864')
    tc.alignment = center_p
    ws_pivot.row_dimensions[1].height = 24

    # Header date
    ws_pivot.cell(row=2, column=1, value='').border = thin_p
    ws_pivot.cell(row=2, column=2, value='Voce').border = thin_p
    ws_pivot.cell(row=2, column=2).font = hdr_font
    ws_pivot.cell(row=2, column=2).fill = hdr_fill
    ws_pivot.cell(row=2, column=2).alignment = center_p
    for ci, (_, label, _) in enumerate(PIVOT_NIGHTS, 3):
        c = ws_pivot.cell(row=2, column=ci, value=label)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center_p; c.border = thin_p
    ws_pivot.row_dimensions[2].height = 18

    # Conta confirmed per notte (no CXL)
    active_rows = [r for r in rows if not r.is_cxl]
    confirmed   = [sum(1 for r in active_rows if getattr(r, nf)) for nf, _, _ in PIVOT_NIGHTS]

    # Carica contratti
    contracts   = {str(c.date): c.rooms for c in
                   HotelContract.query.filter_by(hotel=hotel_name).all()}
    contract    = [contracts.get(str(d), 0) for _, _, d in PIVOT_NIGHTS]
    delta       = [confirmed[i] - contract[i] for i in range(len(PIVOT_NIGHTS))]

    rows_data = [
        ('CONFIRMED', confirmed, PatternFill('solid', start_color='DBEAFE')),
        ('BY CONTRACT', contract, PatternFill('solid', start_color='EDE9FE')),
        ('DELTA', delta, None),
    ]

    for ri, (label, values, bg) in enumerate(rows_data, 3):
        ws_pivot.cell(row=ri, column=1, value='').border = thin_p
        lc = ws_pivot.cell(row=ri, column=2, value=label)
        lc.font = tot_font; lc.border = thin_p; lc.alignment = left_p
        if bg:
            lc.fill = bg
        for ci, val in enumerate(values, 3):
            c = ws_pivot.cell(row=ri, column=ci, value=val if val else None)
            c.border = thin_p; c.alignment = center_p
            c.font = tot_font if label != 'DELTA' else Font(name='Calibri', size=10,
                bold=True,
                color='CC0000' if (val < 0) else ('00AA00' if val > 0 else '888888'))
            if bg:
                c.fill = bg
        ws_pivot.row_dimensions[ri].height = 16

    ws_pivot.column_dimensions['A'].width = 2
    ws_pivot.column_dimensions['B'].width = 14
    for ci in range(3, len(PIVOT_NIGHTS) + 3):
        ws_pivot.column_dimensions[get_column_letter(ci)].width = 10

    # ── Foglio 2: Rooming List ────────────────────────────────────────────────
    ws = wb.create_sheet('Rooming List')

    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', start_color='1F3864')
    norm_font = Font(name='Calibri', size=10, color='000000')
    green_fill  = PatternFill('solid', start_color='C6EFCE')
    yellow_fill = PatternFill('solid', start_color='FFEB9C')
    red_fill    = PatternFill('solid', start_color='FFC7CE')
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    # Riga 1 — Titolo
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(EXPORT_COLS))
    tc = ws.cell(row=1, column=1,
                 value=f'ROOMING LIST — {hotel_name.upper()}')
    tc.font      = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 26

    # Riga 2 — Sottotitolo
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2,   end_column=len(EXPORT_COLS))
    sub = f"Import: {batch_id[:19].replace('_', ' ')}"
    if prev_batch_id:
        sub += (f"  |  Confronto con: {prev_batch_id[:19].replace('_', ' ')}"
                f"  |  🟡 Modificato   🟢 Nuovo   🔴 CXL")
    ws.cell(row=2, column=1, value=sub).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    # Riga 3 — Intestazioni
    for col, (field, label) in enumerate(EXPORT_COLS, 1):
        c = ws.cell(row=3, column=col, value=label)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = center
        c.border    = thin
    ws.row_dimensions[3].height = 20

    def sort_key(r):
        if r.is_cxl:
            group = 2
        elif r.change_date or (r.latest_changes and str(r.latest_changes).strip()):
            group = 1
        else:
            group = 0
        return (group, str(r.last_name or '').upper(), str(r.first_name or '').upper())

    rows_sorted = sorted(rows, key=sort_key)

    for rn, row in enumerate(rows_sorted, 4):
        ref = row.internal_reference or f'__noref_{row.last_name}_{row.first_name}'
        is_cxl     = row.is_cxl
        is_changed = row.change_date or (row.latest_changes and str(row.latest_changes).strip())

        if is_cxl:
            row_fill = red_fill
        elif is_changed:
            row_fill = yellow_fill
        else:
            row_fill = green_fill

        for col, (field, _) in enumerate(EXPORT_COLS, 1):
            c        = ws.cell(row=rn, column=col, value=cell_val(field, row))
            c.border = thin
            c.alignment = left
            c.font = norm_font
            c.fill = row_fill
        ws.row_dimensions[rn].height = 15

    for col, (field, _) in enumerate(EXPORT_COLS, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(field, 14)

    ws.freeze_panes = 'A4'

    # Riepilogo
    total    = len(rows_sorted)
    n_cxl    = sum(1 for r in rows_sorted if r.is_cxl)
    n_active = total - n_cxl
    n_new    = sum(1 for r in rows_sorted
                   if diff.get(r.internal_reference or
                               f'__noref_{r.last_name}_{r.first_name}') == ['NEW'])
    n_chg    = sum(1 for r in rows_sorted
                   if (r.internal_reference or
                       f'__noref_{r.last_name}_{r.first_name}') in diff
                   and diff[r.internal_reference or
                            f'__noref_{r.last_name}_{r.first_name}'] != ['NEW'])

    sr = total + 5
    ws.cell(row=sr, column=1,
            value=(f'Totale: {total}  |  Attivi: {n_active}  |  CXL: {n_cxl}'
                   f'  |  Nuovi: {n_new}  |  Modificati: {n_chg}')
            ).font = Font(name='Calibri', bold=True, size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, total, n_active, n_cxl


# ── Routes ───────────────────────────────────────────────────────────────────

@rooming_bp.route('/')
def index():
    batches = get_batches()
    return render_template('index.html', batches=batches)


@rooming_bp.route('/upload', methods=['POST'])
def upload():
    import re as _re
    from datetime import date as _date

    f = request.files.get('file')
    if not f or not f.filename.endswith('.xlsx'):
        flash('Seleziona un file .xlsx valido.', 'error')
        return redirect(url_for('rooming.index'))

    # Leggi file_timestamp dai metadati del file Excel
    file_ts = None
    try:
        import openpyxl as _opx
        from io import BytesIO as _BytesIO
        raw_bytes = f.read()
        wb_meta = _opx.load_workbook(_BytesIO(raw_bytes), read_only=True)
        props = wb_meta.properties
        file_ts = props.modified or props.created
        wb_meta.close()
        f.seek(0)
    except Exception:
        f.seek(0)

    try:
        df = pd.read_excel(f, sheet_name='RL', header=0)
    except Exception as e:
        flash(f'Errore lettura file: {e}', 'error')
        return redirect(url_for('rooming.index'))

    def parse_latest_changes(text):
        """Estrae (change_type, change_date) da latest_changes."""
        if not text or str(text).strip().lower() in ('nan', ''):
            return None, None
        t = str(text).strip()
        # Pattern: TIPO - DD.MM.YYYY o TIPO - DD/MM/YYYY
        m = _re.match(r'^([A-Za-z\s]+?)\s*[-–]\s*(\d{2}[./]\d{2}[./]\d{4})', t)
        if not m:
            return 'OTHER', None
        raw_type = m.group(1).strip().upper()
        raw_date = m.group(2).replace('/', '.')
        # Normalizza tipo
        if 'CXL' in raw_type:
            ctype = 'CXL'
        elif 'NEW HOTEL' in raw_type:
            ctype = 'NEW HOTEL'
        elif 'NEW BILLING' in raw_type:
            ctype = 'NEW BILLING'
        elif 'NIGHTS' in raw_type or 'NIGHT' in raw_type:
            ctype = 'NIGHTS CHANGE'
        elif 'NAME' in raw_type:
            ctype = 'NAME CHANGE'
        elif 'NEW' in raw_type:
            ctype = 'NEW'
        elif 'OTHER' in raw_type:
            ctype = 'OTHER'
        else:
            ctype = raw_type[:50]
        # Parsa data
        try:
            parts = raw_date.split('.')
            cdate = _date(int(parts[2]), int(parts[1]), int(parts[0]))
        except Exception:
            cdate = None
        return ctype, cdate

    batch_id = make_batch_id(f.filename)
    inserted = 0
    skipped  = 0

    for _, row in df.iterrows():
        first = clean_value('first_name', row.iloc[26] if len(row) > 26 else None)
        last  = clean_value('last_name',  row.iloc[25] if len(row) > 25 else None)
        if not first and not last:
            skipped += 1
            continue

        obj = RoomingList(import_batch=batch_id, file_timestamp=file_ts)
        for col_idx, field in COL_MAP.items():
            raw = row.iloc[col_idx] if col_idx < len(row) else None
            setattr(obj, field, clean_value(field, raw))

        # Deriva change_type e change_date da latest_changes
        ctype, cdate = parse_latest_changes(obj.latest_changes)
        obj.change_type = ctype
        obj.change_date = cdate

        # Categorizza room_category
        is_no_need   = str(obj.night_no_need or '').strip().lower() == 'yes'
        is_spouse    = str(obj.status_spouse or '').strip().lower() == 'yes'
        has_parent   = bool(obj.internal_parent_reference)
        has_hotel    = bool(obj.hotel)
        NFIELDS = ['night_sat_28mar','night_sun_29mar','night_mon_30mar','night_tue_31mar',
                   'night_wed_1apr','night_thu_2apr','night_fri_3apr','night_sat_4apr']
        has_nights   = any(getattr(obj, f) for f in NFIELDS)

        if is_no_need and (has_parent or is_spouse):
            obj.room_category = 'shares_room'
        elif not has_hotel:
            obj.room_category = 'no_hotel'
        elif is_no_need and not has_parent and not is_spouse:
            obj.room_category = 'no_need_solo'
        elif has_hotel and not has_nights and not is_no_need and (obj.registration_state or '').strip().upper() != 'CXL':
            obj.room_category = 'no_dates'
        else:
            obj.room_category = None

        db.session.add(obj)
        inserted += 1

    db.session.commit()

    # ── Secondo passaggio: eredita notti per shares_room ─────────────────────
    from models.models import ManualAssociation

    all_batch = RoomingList.query.filter_by(import_batch=batch_id).all()
    ref_map   = {str(r.internal_reference).strip(): r
                 for r in all_batch if r.internal_reference}

    # Carica associazioni manuali salvate: {(LAST, FIRST): (PARTNER_LAST, PARTNER_FIRST)}
    manual_assoc = {
        (a.last_name.upper(), (a.first_name or '').upper()):
        (a.partner_last_name.upper(), (a.partner_first_name or '').upper())
        for a in ManualAssociation.query.all()
    }
    # Mappa nome → record nel batch corrente
    name_map = {
        ((r.last_name or '').upper(), (r.first_name or '').upper()): r
        for r in all_batch
    }

    NIGHT_FIELDS = ['night_sat_28mar', 'night_sun_29mar', 'night_mon_30mar',
                    'night_tue_31mar', 'night_wed_1apr',  'night_thu_2apr',
                    'night_fri_3apr',  'night_sat_4apr']

    updated = 0
    for r in all_batch:
        if r.room_category != 'shares_room':
            continue
        # Già ha notti proprie — niente da fare
        if any(getattr(r, f) for f in NIGHT_FIELDS):
            continue

        parent = None

        # 1. Via internal_parent_reference
        if r.internal_parent_reference:
            parent = ref_map.get(str(r.internal_parent_reference).strip())

        # 2. Via associazione manuale salvata
        if not parent:
            key = ((r.last_name or '').upper(), (r.first_name or '').upper())
            if key in manual_assoc:
                partner_key = manual_assoc[key]
                parent = name_map.get(partner_key)

        # 3. Via nome nel campo upgrade/comment del batch
        if not parent:
            fname = (r.first_name or '').strip().lower()
            lname = (r.last_name  or '').strip().lower()
            for candidate in all_batch:
                if candidate.id == r.id:
                    continue
                text = ' '.join([
                    str(candidate.upgrade or ''),
                    str(candidate.comment or ''),
                ]).lower()
                if (fname and fname in text) or (lname and lname in text):
                    if any(getattr(candidate, f) for f in NIGHT_FIELDS):
                        parent = candidate
                        break

        if parent and any(getattr(parent, f) for f in NIGHT_FIELDS):
            for f in NIGHT_FIELDS:
                setattr(r, f, getattr(parent, f))
            if not r.hotel and parent.hotel:
                r.hotel = parent.hotel
            updated += 1
        else:
            # Parent non trovato o senza notti — metti in Da controllare
            r.room_category = 'no_need_solo'

    if updated:
        db.session.commit()

    # ── Applica override manuali al nuovo batch ──────────────────────────────
    from models.models import ManualOverride
    overrides = ManualOverride.query.all()
    if overrides:
        batch_rows = RoomingList.query.filter_by(import_batch=batch_id).all()
        ref_to_row = {str(r.internal_reference).strip(): r
                      for r in batch_rows if r.internal_reference}
        applied = 0
        for ov in overrides:
            row = ref_to_row.get(str(ov.internal_reference).strip())
            if row:
                setattr(row, ov.field, ov.override_value)
                applied += 1
        if applied:
            db.session.commit()

    flash(f'Import completato: {inserted} partecipanti caricati, {updated} notti ereditate (batch: {batch_id[:19]}).', 'success')
    return redirect(url_for('rooming.index'))


@rooming_bp.route('/latest-batch')
def latest_batch():
    """Redirect all'ultimo batch importato."""
    batches = get_batches()
    if not batches:
        flash('Nessun batch disponibile.', 'error')
        return redirect(url_for('rooming.index'))
    return redirect(url_for('rooming.batch_detail', batch_id=batches[0][0]))


@rooming_bp.route('/batch/<path:batch_id>')
def batch_detail(batch_id):
    rows = RoomingList.query.filter_by(import_batch=batch_id)\
                            .order_by(RoomingList.hotel, RoomingList.last_name)\
                            .all()
    if not rows:
        flash('Batch non trovato.', 'error')
        return redirect(url_for('rooming.index'))

    hotels = {}
    da_controllare_no_need = []   # night_no_need=Yes senza parent_ref
    da_controllare_no_hotel = []  # hotel vuoto (qualunque notte)

    da_controllare_no_need  = []
    da_controllare_no_hotel = []
    da_controllare_no_dates = []

    for r in rows:
        if r.room_category == 'no_need_solo':
            da_controllare_no_need.append(r)
        elif r.room_category == 'no_hotel':
            da_controllare_no_hotel.append(r)
        elif r.room_category == 'no_dates':
            da_controllare_no_dates.append(r)
        elif not r.hotel or not r.hotel.strip():
            da_controllare_no_hotel.append(r)
        else:
            hotels.setdefault(r.hotel, []).append(r)

    da_controllare = {
        'no_need':  da_controllare_no_need,
        'no_hotel': da_controllare_no_hotel,
        'no_dates': da_controllare_no_dates,
    }

    # Raggruppa modifiche per hotel: {hotel: [(change_type, change_date, count), ...]}
    hotel_changes = {}
    for h, hrows in hotels.items():
        counts = {}
        for r in hrows:
            if r.change_type:
                date_str = r.change_date.strftime('%d/%m') if r.change_date else '?'
                key = (r.change_type, date_str)
                counts[key] = counts.get(key, 0) + 1
        hotel_changes[h] = sorted(
            [(ct, cd, n) for (ct, cd), n in counts.items()],
            key=lambda x: x[1]
        )

    # Batch precedente per diff
    all_batches = get_batches()
    batch_ids   = [b[0] for b in all_batches]
    try:
        idx          = batch_ids.index(batch_id)
        prev_batch   = batch_ids[idx + 1] if idx + 1 < len(batch_ids) else None
    except ValueError:
        prev_batch   = None

    diff = {}
    if prev_batch:
        current_dict  = fetch_batch_dict(batch_id)
        previous_dict = fetch_batch_dict(prev_batch)
        diff          = compute_diff(current_dict, previous_dict)

    return render_template('batch_detail.html',
                           batch_id=batch_id,
                           hotels=hotels,
                           hotel_changes=hotel_changes,
                           da_controllare=da_controllare,
                           diff=diff,
                           prev_batch=prev_batch)


@rooming_bp.route('/export/<path:batch_id>/<hotel_name>')
def export_hotel(batch_id, hotel_name):
    rows = RoomingList.query.filter_by(import_batch=batch_id, hotel=hotel_name).all()
    if not rows:
        flash('Nessun partecipante trovato.', 'error')
        return redirect(url_for('rooming.batch_detail', batch_id=batch_id))

    all_batches = get_batches()
    batch_ids   = [b[0] for b in all_batches]
    try:
        idx        = batch_ids.index(batch_id)
        prev_batch = batch_ids[idx + 1] if idx + 1 < len(batch_ids) else None
    except ValueError:
        prev_batch = None

    diff = {}
    if prev_batch:
        diff = compute_diff(fetch_batch_dict(batch_id), fetch_batch_dict(prev_batch))

    buf, total, active, cxl = build_hotel_excel(hotel_name, rows, diff, batch_id, prev_batch)
    safe = hotel_name.replace('/', '-').replace(' ', '_')
    ts   = datetime.now().strftime('%Y%m%d')
    return send_file(buf,
                     as_attachment=True,
                     download_name=f'RoomingList_{safe}_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rooming_bp.route('/export-da-controllare/<path:batch_id>')
def export_da_controllare(batch_id):
    """Excel con sezioni separate: no-need senza parent + senza hotel."""
    rows = RoomingList.query.filter_by(import_batch=batch_id).all()

    no_need_rows  = sorted(
        [r for r in rows if r.room_category == 'no_need_solo'],
        key=lambda r: (str(r.last_name or '').upper(), str(r.first_name or '').upper())
    )
    no_hotel_rows = sorted(
        [r for r in rows if r.room_category == 'no_hotel' or (not r.hotel and r.room_category not in ('no_need_solo','shares_room','no_dates'))],
        key=lambda r: (str(r.last_name or '').upper(), str(r.first_name or '').upper())
    )
    no_dates_rows = sorted(
        [r for r in rows if r.room_category == 'no_dates'],
        key=lambda r: (str(r.last_name or '').upper(), str(r.first_name or '').upper())
    )

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Da Controllare'

    hdr_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    norm_font  = Font(name='Calibri', size=10)
    sec1_fill  = PatternFill('solid', start_color='FEF3C7')  # giallo — no need
    sec2_fill  = PatternFill('solid', start_color='FEE2E2')  # rosso — senza hotel
    sect_font1 = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    n_cols = len(PREVIEW_COLS)

    # Titolo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value='DA CONTROLLARE')
    tc.font = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ts_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.cell(row=2, column=1,
            value=f'Generato il {ts_str}  |  {len(no_need_rows)} no-need  +  {len(no_hotel_rows)} senza hotel  +  {len(no_dates_rows)} senza date'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    row_num = 3

    def write_section(title, fill_color, rows_list):
        nonlocal row_num
        if not rows_list:
            return
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=n_cols)
        hc = ws.cell(row=row_num, column=1, value=title)
        hc.font = sect_font1
        hc.fill = PatternFill('solid', start_color=fill_color)
        hc.alignment = left
        ws.row_dimensions[row_num].height = 20
        row_num += 1
        for col, (field, label) in enumerate(PREVIEW_COLS, 1):
            c = ws.cell(row=row_num, column=col, value=label)
            c.font = hdr_font
            c.fill = PatternFill('solid', start_color='1F3864')
            c.alignment = center; c.border = thin
        ws.row_dimensions[row_num].height = 18
        row_num += 1
        for r in rows_list:
            for col, (field, _) in enumerate(PREVIEW_COLS, 1):
                c = ws.cell(row=row_num, column=col, value=cell_val(field, r))
                c.font = norm_font
                c.fill = PatternFill('solid', start_color=fill_color)
                c.alignment = left; c.border = thin
            ws.row_dimensions[row_num].height = 15
            row_num += 1

    write_section('NO NEED — senza parent ref', 'FEF3C7', no_need_rows)
    write_section('SENZA HOTEL', 'FEE2E2', no_hotel_rows)
    write_section('SENZA DATE — hotel assegnato ma notti mancanti', 'EDE9FE', no_dates_rows)

    for col, (field, _) in enumerate(PREVIEW_COLS, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(field, 14)
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'DaControllare_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@rooming_bp.route('/export-all/<path:batch_id>')
def export_all(batch_id):
    """Scarica un ZIP con un Excel per ogni hotel."""
    import zipfile

    all_batches = get_batches()
    batch_ids   = [b[0] for b in all_batches]
    try:
        idx        = batch_ids.index(batch_id)
        prev_batch = batch_ids[idx + 1] if idx + 1 < len(batch_ids) else None
    except ValueError:
        prev_batch = None

    diff = {}
    if prev_batch:
        diff = compute_diff(fetch_batch_dict(batch_id), fetch_batch_dict(prev_batch))

    rows  = RoomingList.query.filter_by(import_batch=batch_id).all()
    hotels = {}
    for r in rows:
        h = r.hotel or 'SENZA_HOTEL'
        hotels.setdefault(h, []).append(r)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for hotel, hrows in sorted(hotels.items()):
            buf, _, _, _ = build_hotel_excel(hotel, hrows, diff, batch_id, prev_batch)
            safe = hotel.replace('/', '-').replace(' ', '_')
            ts   = datetime.now().strftime('%Y%m%d')
            zf.writestr(f'RoomingList_{safe}_{ts}.xlsx', buf.read())

    zip_buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(zip_buf,
                     as_attachment=True,
                     download_name=f'RoomingList_ALL_{ts}.zip',
                     mimetype='application/zip')


@rooming_bp.route('/report-category/<path:batch_id>', methods=['POST'])
def report_category(batch_id):
    """Excel con partecipanti filtrati per company_category, raggruppati per hotel."""
    from sqlalchemy import or_
    categories = request.form.getlist('categories')
    if not categories:
        flash('Seleziona almeno una categoria.', 'error')
        return redirect(url_for('rooming.batch_detail', batch_id=batch_id))

    # Gestisce blank separatamente
    conditions = []
    for cat in categories:
        if cat == '__blank__':
            conditions.append(RoomingList.company_category == None)
        else:
            conditions.append(RoomingList.company_category == cat)

    rows = RoomingList.query.filter_by(import_batch=batch_id)\
                            .filter(or_(*conditions))\
                            .order_by(RoomingList.hotel, RoomingList.last_name,
                                      RoomingList.first_name).all()

    # Ordina per gruppo: invariati, modificati, CXL — dentro ogni hotel
    def _cat_sort(r):
        group = 1 if r.is_cxl else 0
        return (r.hotel or '', group,
                str(r.last_name or '').upper(),
                str(r.first_name or '').upper())

    rows = sorted(rows, key=_cat_sort)

    if not rows:
        flash('Nessun partecipante trovato per le categorie selezionate.', 'error')
        return redirect(url_for('rooming.batch_detail', batch_id=batch_id))

    cat_labels = [c if c != '__blank__' else '(blank)' for c in categories]
    title_str  = ' · '.join(cat_labels)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Report'

    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', start_color='1F3864')
    norm_font = Font(name='Calibri', size=10)
    green_fill  = PatternFill('solid', start_color='C6EFCE')
    yellow_fill = PatternFill('solid', start_color='FFEB9C')
    red_fill    = PatternFill('solid', start_color='FFC7CE')
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    n_cols = len(EXPORT_COLS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value=f'REPORT — {title_str.upper()}')
    tc.font = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ts_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.cell(row=2, column=1,
            value=f'Generato il {ts_str}  |  {len(rows)} partecipanti'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    current_hotel = None
    row_num = 3

    for r in rows:
        hotel = r.hotel or 'SENZA HOTEL'
        if hotel != current_hotel:
            current_hotel = hotel
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=n_cols)
            hc = ws.cell(row=row_num, column=1, value=hotel.upper())
            hc.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
            hc.fill = PatternFill('solid', start_color='2F5496')
            hc.alignment = left
            ws.row_dimensions[row_num].height = 20
            row_num += 1
            for col, (field, label) in enumerate(EXPORT_COLS, 1):
                c = ws.cell(row=row_num, column=col, value=label)
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = center; c.border = thin
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        is_cxl    = r.is_cxl
        is_changed = r.change_date or (r.latest_changes and str(r.latest_changes).strip())
        row_fill = red_fill if is_cxl else (yellow_fill if is_changed else green_fill)

        for col, (field, _) in enumerate(EXPORT_COLS, 1):
            c = ws.cell(row=row_num, column=col, value=cell_val(field, r))
            c.font = norm_font; c.fill = row_fill
            c.alignment = left; c.border = thin
        ws.row_dimensions[row_num].height = 15
        row_num += 1

    for col, (field, _) in enumerate(EXPORT_COLS, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(field, 14)
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = title_str.replace(' ', '_').replace('/', '-')[:30]
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'Report_{safe}_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rooming_bp.route('/api/pivot/<path:batch_id>')
def api_pivot(batch_id):
    """Ritorna JSON con pivot notti x hotel — esclude solo CXL."""
    rows = RoomingList.query.filter_by(import_batch=batch_id)\
                            .filter(RoomingList.registration_state != 'CXL')\
                            .all()
    if not rows:
        return jsonify({'error': 'Batch non trovato'}), 404

    NIGHTS = [
        ('night_no_need',    'No need'),
        ('night_sat_28mar',  'Sat 28/03'),
        ('night_sun_29mar',  'Sun 29/03'),
        ('night_mon_30mar',  'Mon 30/03'),
        ('night_tue_31mar',  'Tue 31/03 (Nexus)'),
        ('night_wed_1apr',   'Wed 01/04 (Nexus)'),
        ('night_thu_2apr',   'Thu 02/04'),
        ('night_fri_3apr',   'Fri 03/04'),
        ('night_sat_4apr',   'Sat 04/04'),
        ('night_other',      'Other…'),
    ]

    hotels = sorted(set(r.hotel or 'None' for r in rows))
    pivot  = {nf: {h: 0 for h in hotels} for nf, _ in NIGHTS}

    for r in rows:
        h = r.hotel or 'None'
        for nf, _ in NIGHTS:
            if getattr(r, nf):
                pivot[nf][h] += 1

    return jsonify({
        'hotels': hotels,
        'nights': [{'field': nf, 'label': nl} for nf, nl in NIGHTS],
        'pivot':  {nf: pivot[nf] for nf, _ in NIGHTS},
    })



@rooming_bp.route('/export-pivot/<path:batch_id>')
def export_pivot(batch_id):
    rows = RoomingList.query.filter_by(import_batch=batch_id)\
                            .filter(RoomingList.registration_state != 'CXL')\
                            .all()
    if not rows:
        flash('Batch non trovato.', 'error')
        return redirect(url_for('rooming.index'))

    # Definizione notti in ordine
    NIGHTS = [
        ('night_no_need',    'No need'),
        ('night_sat_28mar',  'Night of Saturday, March 28th'),
        ('night_sun_29mar',  'Night of Sunday, March 29th'),
        ('night_mon_30mar',  'Night of Monday, March 30th'),
        ('night_tue_31mar',  'Night of Tuesday, March 31st (covered by Nexus)'),
        ('night_wed_1apr',   'Night of Wednesday, April 1st (covered by Nexus)'),
        ('night_thu_2apr',   'Night of Thursday, April 2nd'),
        ('night_fri_3apr',   'Night of Friday, April 3rd'),
        ('night_sat_4apr',   'Night of Saturday, April 4th'),
        ('night_other',      'Other…'),
    ]

    # Hotel in ordine alfabetico
    hotels = sorted(set(r.hotel or 'None' for r in rows))

    # Costruisce matrice pivot[notte][hotel] = count
    pivot = {nf: {h: 0 for h in hotels} for nf, _ in NIGHTS}
    for r in rows:
        h = r.hotel or 'None'
        for nf, _ in NIGHTS:
            if getattr(r, nf):
                pivot[nf][h] += 1

    # ── Excel ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pivot Notti x Hotel'

    hdr_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill   = PatternFill('solid', start_color='1F3864')
    tot_font   = Font(name='Calibri', bold=True, size=10)
    tot_fill   = PatternFill('solid', start_color='D9E1F2')
    norm_font  = Font(name='Calibri', size=10)
    zero_font  = Font(name='Calibri', size=10, color='CCCCCC')
    center     = Alignment(horizontal='center', vertical='center')
    left       = Alignment(horizontal='left',   vertical='center')
    thin       = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'),  bottom=Side(style='thin'))

    # Riga 1 — titolo
    n_cols = len(hotels) + 2  # notte + hotels + grand total
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value='PIVOT — NOTTI PER HOTEL')
    tc.font = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 26

    # Riga 2 — sottotitolo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ts_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.cell(row=2, column=1,
            value=f'Generato il {ts_str}  |  batch: {batch_id[:19]}'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    # Riga 3 — intestazioni colonne
    ws.cell(row=3, column=1, value='Notte').font = hdr_font
    ws.cell(row=3, column=1).fill = hdr_fill
    ws.cell(row=3, column=1).alignment = left
    ws.cell(row=3, column=1).border = thin

    for ci, h in enumerate(hotels, 2):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = thin

    # Colonna Grand Total
    gt_col = len(hotels) + 2
    c = ws.cell(row=3, column=gt_col, value='Grand Total')
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = center; c.border = thin
    ws.row_dimensions[3].height = 20

    # Righe dati
    for rn, (nf, label) in enumerate(NIGHTS, 4):
        ws.cell(row=rn, column=1, value=label).font = norm_font
        ws.cell(row=rn, column=1).border = thin
        ws.cell(row=rn, column=1).alignment = left

        row_total = 0
        for ci, h in enumerate(hotels, 2):
            val = pivot[nf][h]
            row_total += val
            c = ws.cell(row=rn, column=ci, value=val if val else None)
            c.border = thin; c.alignment = center
            c.font = norm_font if val else zero_font

        # Grand Total riga
        c = ws.cell(row=rn, column=gt_col, value=row_total if row_total else None)
        c.font = tot_font; c.fill = tot_fill
        c.border = thin; c.alignment = center
        ws.row_dimensions[rn].height = 16

    # Riga Grand Total colonne
    tot_row = len(NIGHTS) + 4
    ws.cell(row=tot_row, column=1, value='Grand Total').font = tot_font
    ws.cell(row=tot_row, column=1).fill = tot_fill
    ws.cell(row=tot_row, column=1).border = thin

    grand_total = 0
    for ci, h in enumerate(hotels, 2):
        col_total = sum(pivot[nf][h] for nf, _ in NIGHTS)
        grand_total += col_total
        c = ws.cell(row=tot_row, column=ci, value=col_total if col_total else None)
        c.font = tot_font; c.fill = tot_fill
        c.border = thin; c.alignment = center

    c = ws.cell(row=tot_row, column=gt_col, value=grand_total)
    c.font = Font(name='Calibri', bold=True, size=11, color='1F3864')
    c.fill = PatternFill('solid', start_color='B8CCE4')
    c.border = thin; c.alignment = center
    ws.row_dimensions[tot_row].height = 18

    # Larghezze colonne
    ws.column_dimensions['A'].width = 44
    for ci in range(2, gt_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    ws.freeze_panes = 'B4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'Pivot_Notti_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')




@rooming_bp.route('/batch/delete/<path:batch_id>', methods=['POST'])
def delete_batch(batch_id):
    RoomingList.query.filter_by(import_batch=batch_id).delete()
    db.session.commit()
    flash('Batch eliminato.', 'success')
    return redirect(url_for('rooming.index'))


# ── Export dinamico ──────────────────────────────────────────────────────────

# Tutte le colonne disponibili con etichetta
ALL_COLS = [
    ('title',                'Title'),
    ('last_name',            'Last Name'),
    ('first_name',           'First Name'),
    ('comment',              'Comment'),
    ('email',                'Email'),
    ('phone',                'Phone'),
    ('billing',              'Billing'),
    ('upgrade',              'Upgrade'),
    ('check_in',             'Check In'),
    ('check_out',            'Check Out'),
    ('company_name',         'Company'),
    ('company_country',      'Country'),
    ('job_position',         'Job Position'),
    ('email',                'Email'),
    ('phone',                'Phone'),
    ('company_category',     'Category'),
    ('company_subcategory',  'Sub-Category'),
    ('registration_state',   'Reg. State'),
    ('hotel',                'Hotel'),
    ('upgrade',              'Upgrade'),
    ('billing',              'Billing'),
    ('nexus_bd',             'NEXUS BD'),
    ('night_no_need',        'No need'),
    ('night_sat_28mar',      'Sat 28/03'),
    ('night_sun_29mar',      'Sun 29/03'),
    ('night_mon_30mar',      'Mon 30/03'),
    ('night_tue_31mar',      'Tue 31/03'),
    ('night_wed_1apr',       'Wed 01/04'),
    ('night_thu_2apr',       'Thu 02/04'),
    ('night_fri_3apr',       'Fri 03/04'),
    ('night_sat_4apr',       'Sat 04/04'),
    ('arrival_mode',         'Arrival'),
    ('need_smooth_checkin',  'Smooth CI'),
    ('diet_restrictions',    'Diet'),
    ('need_visa',            'Visa?'),
    ('visa_birth_date',      'Birth Date'),
    ('visa_birth_place',     'Birth Place'),
    ('visa_passport',        'Passport N.'),
    ('visa_delivery_date',   'Doc Issued'),
    ('visa_expiration_date', 'Doc Expiry'),
    ('visa_company_address', 'Company Address'),
    ('comment',              'Comment'),
    ('internal_reference',   'Internal Ref'),
    ('participant_number',   'Participant N.'),
    ('ean8_barcode',         'EAN8'),
    ('status_vp_bd',         'VP/BD'),
    ('status_organisator',   'Organisator'),
    ('status_board_nai',     'Board NAI'),
    ('status_climate_day',   'Climate Day'),
    ('status_spouse',        'Spouse Flag'),
    ('spouse_name',          'Spouse Name'),   # colonna virtuale — valorizzata solo se elaborate_spouse
    ('is_parent_manager',    'Parent Mgr'),
    ('registered_colleagues','Colleagues'),
    ('latest_changes',       'Latest Changes'),
    ('change_type',          'Change Type'),
    ('change_date',          'Change Date'),
    ('file_timestamp',       'File Timestamp'),
]

@rooming_bp.route('/export-custom/<path:batch_id>')
def export_custom(batch_id):
    from models.models import ExportConfig
    rows = RoomingList.query.filter_by(import_batch=batch_id).all()
    if not rows:
        flash('Batch non trovato.', 'error')
        return redirect(url_for('rooming.index'))

    hotels       = sorted(set(r.hotel or 'SENZA HOTEL' for r in rows))
    stati        = sorted(set(r.registration_state or '' for r in rows if r.registration_state))
    saved_configs = [c.name for c in ExportConfig.query.order_by(ExportConfig.name).all()]

    return render_template('export_custom.html',
                           batch_id=batch_id,
                           all_cols=ALL_COLS,
                           hotels=hotels,
                           stati=stati,
                           saved_configs=saved_configs)


@rooming_bp.route('/api/preview/<path:batch_id>', methods=['POST'])
def api_preview(batch_id):
    """Ritorna JSON con anteprima righe filtrate e colonne scelte."""
    data        = request.get_json()
    cols        = data.get('cols', [])          # lista di field names in ordine
    f_hotels    = data.get('hotels', [])        # [] = tutti
    f_stati     = data.get('stati', [])         # [] = tutti
    f_notti     = data.get('notti', [])         # [] = tutte
    include_cxl = data.get('include_cxl', True)
    page        = int(data.get('page', 1))
    per_page    = 50

    if not cols:
        return jsonify({'rows': [], 'total': 0, 'cols': []})

    q = RoomingList.query.filter_by(import_batch=batch_id)
    if f_hotels:
        q = q.filter(RoomingList.hotel.in_(f_hotels))
    if f_stati:
        q = q.filter(RoomingList.registration_state.in_(f_stati))
    if not include_cxl:
        q = q.filter(RoomingList.registration_state != 'CXL')

    all_rows = q.order_by(RoomingList.last_name, RoomingList.first_name).all()

    # Filtro notti (AND: deve avere almeno una delle notti selezionate)
    if f_notti:
        night_map = {
            'sat_28': 'night_sat_28mar', 'sun_29': 'night_sun_29mar',
            'mon_30': 'night_mon_30mar', 'tue_31': 'night_tue_31mar',
            'wed_1':  'night_wed_1apr',  'thu_2':  'night_thu_2apr',
            'fri_3':  'night_fri_3apr',  'sat_4':  'night_sat_4apr',
        }
        filtered = []
        for r in all_rows:
            for n in f_notti:
                field = night_map.get(n)
                if field and getattr(r, field):
                    filtered.append(r)
                    break
        all_rows = filtered

    total  = len(all_rows)
    offset = (page - 1) * per_page
    paged  = all_rows[offset:offset + per_page]

    # Costruisce label mappa
    col_labels = {f: l for f, l in ALL_COLS}

    result_rows = []
    for r in paged:
        row_data = {}
        for f in cols:
            row_data[f] = cell_val(f, r)
        result_rows.append(row_data)

    return jsonify({
        'rows':   result_rows,
        'total':  total,
        'page':   page,
        'pages':  (total + per_page - 1) // per_page,
        'labels': {f: col_labels.get(f, f) for f in cols},
    })


@rooming_bp.route('/api/export-custom/<path:batch_id>', methods=['POST'])
def api_export_custom(batch_id):
    """Genera Excel con configurazione dinamica."""
    data              = request.get_json()
    cols              = data.get('cols', [])
    f_hotels          = data.get('hotels', [])
    f_stati           = data.get('stati', [])
    f_notti           = data.get('notti', [])
    include_cxl       = data.get('include_cxl', True)
    title_name        = data.get('title', 'Export')
    elaborate_spouse  = data.get('elaborate_spouse', False)

    if not cols:
        return jsonify({'error': 'Nessuna colonna selezionata'}), 400

    q = RoomingList.query.filter_by(import_batch=batch_id)
    if f_hotels:
        q = q.filter(RoomingList.hotel.in_(f_hotels))
    if f_stati:
        q = q.filter(RoomingList.registration_state.in_(f_stati))
    if not include_cxl:
        q = q.filter(RoomingList.registration_state != 'CXL')

    all_rows = q.order_by(RoomingList.last_name, RoomingList.first_name).all()

    if f_notti:
        night_map = {
            'sat_28': 'night_sat_28mar', 'sun_29': 'night_sun_29mar',
            'mon_30': 'night_mon_30mar', 'tue_31': 'night_tue_31mar',
            'wed_1':  'night_wed_1apr',  'thu_2':  'night_thu_2apr',
            'fri_3':  'night_fri_3apr',  'sat_4':  'night_sat_4apr',
        }
        filtered = []
        for r in all_rows:
            for n in f_notti:
                field = night_map.get(n)
                if field and getattr(r, field):
                    filtered.append(r)
                    break
        all_rows = filtered

    # Ordinamento: prima invariati, poi modificati, in fondo CXL
    def _sort_key(r):
        if r.is_cxl:
            group = 2
        elif r.change_date or (r.latest_changes and str(r.latest_changes).strip()):
            group = 1
        else:
            group = 0
        return (group, str(r.last_name or '').upper(), str(r.first_name or '').upper())

    all_rows = sorted(all_rows, key=_sort_key)

    # ── Elaborazione spouse runtime ───────────────────────────────────────
    spouse_notes  = {}   # id riga spouse -> "Spouse of <nome>"
    spouse_ids    = set()  # id righe che sono spouse
    parent_spouse = {}   # id riga titolare -> nome spouse

    if elaborate_spouse:
        all_batch = RoomingList.query.filter_by(import_batch=batch_id).all()
        ref_to_row = {
            r.internal_reference: r
            for r in all_batch if r.internal_reference
        }
        import re as _re
        for r in all_batch:
            if (r.status_spouse or '').strip().lower() == 'yes':
                spouse_ids.add(r.id)
                spouse_full = f"{r.first_name or ''} {r.last_name or ''}".strip()
                parent_name = None
                parent_row  = None

                # 1. Cerca via internal_parent_reference
                if r.internal_parent_reference:
                    parent_row = ref_to_row.get(r.internal_parent_reference)
                    if parent_row:
                        parent_name = f"{parent_row.first_name or ''} {parent_row.last_name or ''}".strip()

                # 2. Fallback: cerca nel comment pattern noti
                if not parent_name and r.comment:
                    m = _re.search(
                        r'(?:wife|spouse)\s+of\s+(.+)|'
                        r'double\s+room\s+with\s+(.+)|'
                        r'shared\s+room\s+with\s+(.+)|'
                        r'twin\s+room\s+with\s+(.+)',
                        r.comment, _re.IGNORECASE)
                    if m:
                        parent_name = next(g for g in m.groups() if g).strip().rstrip('.')

                # Nota sulla riga spouse
                if parent_name:
                    spouse_notes[r.id] = f"Spouse of {parent_name}"

                # Nota sulla riga titolare
                if parent_row and spouse_full:
                    parent_spouse[parent_row.id] = spouse_full

    col_labels = {f: l for f, l in ALL_COLS}

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Export'

    hdr_font    = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill    = PatternFill('solid', start_color='1F3864')
    norm_font   = Font(name='Calibri', size=10)
    cxl_font    = Font(name='Calibri', size=10, color='CC0000', strike=True)
    spouse_font = Font(name='Calibri', size=10, color='888888', strike=True)
    left        = Alignment(horizontal='left', vertical='center')
    center      = Alignment(horizontal='center', vertical='center')
    thin        = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'),  bottom=Side(style='thin'))

    # Titolo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    tc = ws.cell(row=1, column=1, value=title_name.upper())
    tc.font = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 26

    # Sottotitolo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    ts_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.cell(row=2, column=1,
            value=f'Generato il {ts_str}  |  {len(all_rows)} partecipanti'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    # Header
    for ci, f in enumerate(cols, 1):
        c = ws.cell(row=3, column=ci, value=col_labels.get(f, f))
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = thin
    ws.row_dimensions[3].height = 20

    for rn, row in enumerate(all_rows, 4):
        is_cxl    = row.is_cxl
        is_spouse = row.id in spouse_ids

        for ci, f in enumerate(cols, 1):
            # Colonne virtuali elaborate a runtime
            if f == 'spouse_name':
                val = parent_spouse.get(row.id, '')
            elif is_spouse and f == 'latest_changes' and row.id in spouse_notes:
                val = spouse_notes[row.id]
            else:
                val = cell_val(f, row)

            c = ws.cell(row=rn, column=ci, value=val)
            c.border = thin; c.alignment = left

            if is_cxl:
                c.font = cxl_font
            elif is_spouse:
                # Latest Changes leggibile (non barrato), resto grigio barrato
                if f == 'latest_changes' and row.id in spouse_notes:
                    c.font = Font(name='Calibri', size=10, bold=True, color='1F3864')
                else:
                    c.font = spouse_font
            else:
                c.font = norm_font
        ws.row_dimensions[rn].height = 15

    for ci, f in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(f, 16)

    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'Export_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



# ── Preview report API (unified) ─────────────────────────────────────────────

PREVIEW_COLS = [
    # Sequenza standard (uguale a EXPORT_COLS)
    ('title',                     'Title'),
    ('last_name',                 'Last Name'),
    ('first_name',                'First Name'),
    ('comment',                   'Comment'),
    ('email',                     'Email'),
    ('phone',                     'Phone'),
    ('billing',                   'Billing'),
    ('upgrade',                   'Upgrade'),
    ('arrival_mode',              'Arrival Mode'),
    ('check_in',                  'Check In'),
    ('check_out',                 'Check Out'),
    ('need_visa',                 'Visa?'),
    ('visa_birth_date',           'Birth Date'),
    ('visa_birth_place',          'Birth Place'),
    ('visa_passport',             'Passport N.'),
    ('visa_delivery_date',        'Doc Issued'),
    ('visa_expiration_date',      'Doc Expiry'),
    ('latest_changes',            'Latest Changes'),
    ('change_type',               'Change Type'),
    ('change_date',               'Change Date'),
    # Campi aggiuntivi in coda
    ('registration_state',        'Stato'),
    ('hotel',                     'Hotel'),
    ('company_name',              'Azienda'),
    ('company_country',           'Paese'),
    ('company_category',          'Categoria'),
    ('company_subcategory',       'Sub-Categoria'),
    ('job_position',              'Job Position'),
    ('nexus_bd',                  'NEXUS BD'),
    ('need_smooth_checkin',       'Smooth CI'),
    ('diet_restrictions',         'Dieta'),
    ('internal_reference',        'Internal Ref'),
    ('participant_number',        'N. Partecipante'),
    ('visa_company_address',      'Indirizzo azienda'),
    ('status_vp_bd',              'VP/BD'),
    ('status_organisator',        'Organisator'),
    ('status_board_nai',          'Board NAI'),
    ('status_climate_day',        'Climate Day'),
    ('status_prospective_council','Prospective Council'),
    ('status_spouse',             'Spouse'),
    ('night_sat_28mar',           'Notte 28/03'),
    ('night_sun_29mar',           'Notte 29/03'),
    ('night_mon_30mar',           'Notte 30/03'),
    ('night_tue_31mar',           'Notte 31/03'),
    ('night_wed_1apr',            'Notte 01/04'),
    ('night_thu_2apr',            'Notte 02/04'),
    ('night_fri_3apr',            'Notte 03/04'),
    ('night_sat_4apr',            'Notte 04/04'),
    ('delegation_key',            'Delegation Key'),
    ('internal_parent_reference', 'Parent Ref'),
    ('prospective_response',      'Prospective Response'),
]

@rooming_bp.route('/api/preview-report/<path:batch_id>', methods=['POST'])
def api_preview_report(batch_id):
    from sqlalchemy import or_, and_
    data     = request.get_json()
    rtype    = data.get('type')
    values   = data.get('values', [])
    conditions_raw = data.get('conditions', [])
    logic    = data.get('logic', 'AND')

    q = RoomingList.query.filter_by(import_batch=batch_id)

    if rtype == 'category':
        conds = [RoomingList.company_category == None if v == '__blank__'
                 else RoomingList.company_category == v for v in values]
        q = q.filter(or_(*conds))
    elif rtype == 'billing':
        conds = [RoomingList.billing == None if v == '__blank__'
                 else RoomingList.billing == v for v in values]
        q = q.filter(or_(*conds))
    elif rtype == 'arrival':
        conds = [RoomingList.arrival_mode == None if v == '__blank__'
                 else RoomingList.arrival_mode == v for v in values]
        q = q.filter(or_(*conds))
    elif rtype == 'query':
        conds = []
        for c in conditions_raw:
            field = c.get('field', '')
            op    = c.get('op', 'eq')
            val   = c.get('val', '')
            col   = getattr(RoomingList, field, None)
            if col is None:
                continue
            if op == 'eq':
                conds.append(col == val)
            elif op == 'neq':
                conds.append(col != val)
            elif op == 'contains':
                conds.append(col.ilike(f'%{val}%'))
            elif op == 'not_empty':
                conds.append(col != None)
            elif op == 'empty':
                conds.append(col == None)
        if conds:
            q = q.filter(and_(*conds) if logic == 'AND' else or_(*conds))

    rows = q.order_by(RoomingList.hotel, RoomingList.last_name,
                      RoomingList.first_name).all()

    def _sort(r):
        return (r.hotel or '', 1 if r.is_cxl else 0,
                str(r.last_name or '').upper())
    rows = sorted(rows, key=_sort)

    result = []
    for r in rows:
        row_data = {}
        for field, label in PREVIEW_COLS:
            row_data[label] = cell_val(field, r)
        result.append(row_data)

    return jsonify({'rows': result, 'total': len(result)})


@rooming_bp.route('/report-query/<path:batch_id>', methods=['POST'])
def report_query(batch_id):
    """Excel da query libera."""
    from sqlalchemy import or_, and_
    import json as _json
    conditions_raw = _json.loads(request.form.get('conditions', '[]'))
    logic          = request.form.get('logic', 'AND')

    q = RoomingList.query.filter_by(import_batch=batch_id)
    conds = []
    for c in conditions_raw:
        field = c.get('field', '')
        op    = c.get('op', 'eq')
        val   = c.get('val', '')
        col   = getattr(RoomingList, field, None)
        if col is None:
            continue
        if op == 'eq':       conds.append(col == val)
        elif op == 'neq':    conds.append(col != val)
        elif op == 'contains': conds.append(col.ilike(f'%{val}%'))
        elif op == 'not_empty': conds.append(col != None)
        elif op == 'empty':  conds.append(col == None)
    if conds:
        q = q.filter(and_(*conds) if logic == 'AND' else or_(*conds))

    rows = q.order_by(RoomingList.hotel, RoomingList.last_name,
                      RoomingList.first_name).all()

    def _sort(r):
        return (r.hotel or '', 1 if r.is_cxl else 0,
                str(r.last_name or '').upper(), str(r.first_name or '').upper())
    rows = sorted(rows, key=_sort)

    if not rows:
        flash('Nessun risultato.', 'error')
        return redirect(url_for('rooming.batch_detail', batch_id=batch_id))

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Query Result'

    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', start_color='1F3864')
    norm_font = Font(name='Calibri', size=10)
    green_fill  = PatternFill('solid', start_color='C6EFCE')
    yellow_fill = PatternFill('solid', start_color='FFEB9C')
    red_fill    = PatternFill('solid', start_color='FFC7CE')
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    n_cols = len(PREVIEW_COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value='QUERY RESULT')
    tc.font = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ts_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.cell(row=2, column=1,
            value=f'Generato il {ts_str}  |  {len(rows)} partecipanti'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    current_hotel = None
    row_num = 3
    for r in rows:
        hotel = r.hotel or 'SENZA HOTEL'
        if hotel != current_hotel:
            current_hotel = hotel
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=n_cols)
            hc = ws.cell(row=row_num, column=1, value=hotel.upper())
            hc.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
            hc.fill = PatternFill('solid', start_color='2F5496')
            hc.alignment = left
            ws.row_dimensions[row_num].height = 20
            row_num += 1
            for col, (field, label) in enumerate(PREVIEW_COLS, 1):
                c = ws.cell(row=row_num, column=col, value=label)
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = center; c.border = thin
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        is_changed = r.change_date or (r.latest_changes and str(r.latest_changes).strip())
        row_fill = red_fill if r.is_cxl else (yellow_fill if is_changed else green_fill)
        for col, (field, _) in enumerate(PREVIEW_COLS, 1):
            c = ws.cell(row=row_num, column=col, value=cell_val(field, r))
            c.font = norm_font; c.fill = row_fill
            c.alignment = left; c.border = thin
        ws.row_dimensions[row_num].height = 15
        row_num += 1

    for col, (field, _) in enumerate(PREVIEW_COLS, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(field, 14)
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'Query_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rooming_bp.route('/api/save-config', methods=['POST'])
def api_save_config():
    import json
    from models.models import ExportConfig
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Nome configurazione mancante'}), 400
    cfg = ExportConfig.query.filter_by(name=name).first()
    if not cfg:
        cfg = ExportConfig(name=name)
        db.session.add(cfg)
    cfg.cols        = json.dumps(data.get('cols', []))
    cfg.hotels      = json.dumps(data.get('hotels', []))
    cfg.stati       = json.dumps(data.get('stati', []))
    cfg.notti       = json.dumps(data.get('notti', []))
    cfg.include_cxl = data.get('include_cxl', True)
    db.session.commit()
    all_names = [c.name for c in ExportConfig.query.order_by(ExportConfig.name).all()]
    return jsonify({'ok': True, 'configs': all_names})


@rooming_bp.route('/api/load-config/<path:name>')
def api_load_config(name):
    import json
    from models.models import ExportConfig
    cfg = ExportConfig.query.filter_by(name=name).first()
    if not cfg:
        return jsonify({'error': 'Configurazione non trovata'}), 404
    return jsonify({
        'cols':        json.loads(cfg.cols or '[]'),
        'hotels':      json.loads(cfg.hotels or '[]'),
        'stati':       json.loads(cfg.stati or '[]'),
        'notti':       json.loads(cfg.notti or '[]'),
        'include_cxl': cfg.include_cxl,
    })

# ── Hotel contracts ───────────────────────────────────────────────────────────

# Mapping nomi file → nomi DB
HOTEL_NAME_MAP = {
    'movenpick':    'Movenpick',
    'beau rivage':  'Beau-Rivage Palace',
    'beau-rivage':  'Beau-Rivage Palace',
    'de la paix':   'Hotel de la Paix',
    'lausanne':     'Lausanne Palace',
    'novotel':      'Novotel',
    'royal savoy':  'Royal Savoy',
    'starling':     'Starling',
}

DATES = [
    '2026-03-28', '2026-03-29', '2026-03-30', '2026-03-31',
    '2026-04-01', '2026-04-02', '2026-04-03', '2026-04-04',
]

DATE_LABELS = ['28-mar','29-mar','30-mar','31-mar','1-apr','2-apr','3-apr','4-apr']


def normalize_hotel(raw):
    """Normalizza nome hotel grezzo verso nome DB."""
    s = str(raw).strip().lower()
    for key, val in HOTEL_NAME_MAP.items():
        if key in s:
            return val
    return None





@rooming_bp.route('/upload-activities', methods=['POST'])
def upload_activities():
    """Import file attivita: ogni foglio = attivita.
    Colonne: Last Name, First Name, Composite (opzionale, contiene email tra <>)"""
    from models.models import Activity, ActivityParticipant
    import re as _re2

    f = request.files.get('file')
    if not f or not f.filename.endswith('.xlsx'):
        flash('Seleziona un file .xlsx valido.', 'error')
        return redirect(url_for('rooming.index'))

    batch_id    = request.form.get('batch_id', '')
    activity_id = request.form.get('activity_id', '')

    try:
        xl = pd.ExcelFile(f)
    except Exception as e:
        flash(f'Errore lettura file: {e}', 'error')
        return redirect(url_for('rooming.index'))

    email_map   = {}
    rooming_map = {}
    if batch_id:
        for r in RoomingList.query.filter_by(import_batch=batch_id).all():
            if r.email:
                email_map[r.email.strip().lower()] = r.id
            if r.last_name:
                key = ((r.last_name or '').strip().upper(),
                       (r.first_name or '').strip().upper())
                rooming_map[key] = r.id

    def find_rooming_id(last_name, first_name, composite):
        if composite:
            m = _re2.search(r'<([^>]+)>', composite)
            if m:
                email = m.group(1).strip().lower()
                if email in email_map:
                    return email_map[email]
        key = (last_name.upper(), first_name.upper())
        if key in rooming_map:
            return rooming_map[key]
        for (ln, fn), rid in rooming_map.items():
            if ln == last_name.upper():
                return rid
        return None

    imported_activities = 0
    imported_participants = 0

    for sheet_name in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet_name, header=None)

        last_col = first_col = composite_col = None
        data_start = 0

        for i in range(min(10, len(df))):
            row_vals = [str(df.iloc[i, c]).strip().lower().replace(' ', '').replace('\n', '')
                        for c in range(len(df.columns))]
            for ci, val in enumerate(row_vals):
                if val in ('lastname', 'cognome', 'lastlname'):    last_col = ci
                if val in ('firstname', 'nome', 'firstname'):      first_col = ci
                if val in ('composite', 'email', 'compositename'): composite_col = ci
            if last_col is not None and first_col is not None:
                data_start = i + 1
                break

        if last_col is None or first_col is None:
            continue

        if activity_id:
            activity = Activity.query.get(int(activity_id))
            if not activity:
                flash('Attivita non trovata.', 'error')
                return redirect(url_for('rooming.activities_page'))
            ActivityParticipant.query.filter_by(activity_id=activity.id).delete()
            db.session.flush()
        else:
            date_match = _re2.search(r'(\d{2}\.\d{2}\.\d{4})', sheet_name)
            activity_date = None
            activity_name = sheet_name.strip()
            if date_match:
                try:
                    from datetime import datetime as _dt
                    activity_date = _dt.strptime(date_match.group(1), '%d.%m.%Y').date()
                    activity_name = sheet_name[date_match.end():].strip(' -').strip()
                except Exception:
                    pass
            existing = Activity.query.filter_by(name=activity_name, date=activity_date).first()
            if existing:
                ActivityParticipant.query.filter_by(activity_id=existing.id).delete()
                db.session.flush()
                activity = existing
            else:
                activity = Activity(name=activity_name, date=activity_date, import_batch=batch_id)
                db.session.add(activity)
                db.session.flush()

        for i in range(data_start, len(df)):
            row = df.iloc[i]
            last_name  = str(row[last_col]).strip()  if not pd.isna(row[last_col])  else ''
            first_name = str(row[first_col]).strip() if not pd.isna(row[first_col]) else ''
            composite  = str(row[composite_col]).strip() if composite_col is not None and not pd.isna(row[composite_col]) else ''

            if not last_name or last_name.lower() == 'nan':
                continue

            rooming_id = find_rooming_id(last_name, first_name, composite)

            ap = ActivityParticipant(
                activity_id=activity.id,
                rooming_list_id=rooming_id,
                last_name=last_name,
                first_name=first_name,
            )
            db.session.add(ap)
            imported_participants += 1

        imported_activities += 1

    db.session.commit()
    flash(f'Import: {imported_activities} attivita, {imported_participants} partecipanti.', 'success')
    if activity_id:
        return redirect(url_for('rooming.activities_page'))
    return redirect(url_for('rooming.index'))


@rooming_bp.route('/api/activity-rename/<int:activity_id>', methods=['POST'])
def api_activity_rename(activity_id):
    from models.models import Activity
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'Nome vuoto'})
    a = Activity.query.get(activity_id)
    if not a:
        return jsonify({'ok': False, 'error': 'Non trovata'})
    a.name = name
    db.session.commit()
    return jsonify({'ok': True})


@rooming_bp.route('/export-all-activities')
def export_all_activities():
    """Excel con tutti i partecipanti di tutte le attività."""
    from models.models import Activity, ActivityParticipant

    activities = Activity.query.order_by(Activity.date, Activity.name).all()

    # Colonne per export attività — sequenza ridotta
    ALL_ACT_COLS = [
        ('title',              'Title'),
        ('last_name',          'Last Name'),
        ('first_name',         'First Name'),
        ('comment',            'Comment'),
        ('email',              'Email'),
        ('phone',              'Phone'),
        ('billing',            'Billing'),
        ('company_category',   'Category'),
        ('upgrade',            'Upgrade'),
        ('arrival_mode',       'Arrival Mode'),
        ('check_in',           'Check In'),
        ('check_out',          'Check Out'),
        ('company_name',       'Company'),
        ('company_country',    'Country'),
        ('job_position',       'Job Position'),
        ('company_subcategory','Sub Category'),
        ('diet_restrictions',  'Diet Restriction'),
        ('in_rooming',         'In Rooming List'),
    ]

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'All Activities'

    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', start_color='1F3864')
    norm_font = Font(name='Calibri', size=10)
    warn_font = Font(name='Calibri', size=10, color='CC0000')
    act_font  = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    green_fill  = PatternFill('solid', start_color='C6EFCE')
    yellow_fill = PatternFill('solid', start_color='FFEB9C')
    red_fill    = PatternFill('solid', start_color='FFC7CE')
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    n_cols  = len(ALL_ACT_COLS)
    row_num = 1

    # Titolo
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=n_cols)
    tc = ws.cell(row=row_num, column=1, value='TUTTI I PARTECIPANTI — TUTTE LE ATTIVITÀ')
    tc.font = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[row_num].height = 26
    row_num += 1

    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=n_cols)
    ws.cell(row=row_num, column=1,
            value=f'Generato il {datetime.now().strftime("%d/%m/%Y %H:%M")}'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[row_num].height = 14
    row_num += 1

    for activity in activities:
        parts = sorted(activity.participants,
                       key=lambda x: (x.last_name or '').upper())
        if not parts:
            continue

        # Intestazione attività
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=n_cols)
        hc = ws.cell(row=row_num, column=1,
                     value=f'{activity.name.upper()}  —  {activity.date.strftime("%d/%m/%Y") if activity.date else ""}')
        hc.font = act_font
        hc.fill = PatternFill('solid', start_color='2F5496')
        hc.alignment = left
        ws.row_dimensions[row_num].height = 20
        row_num += 1

        # Header colonne
        for col, (field, label) in enumerate(ALL_ACT_COLS, 1):
            c = ws.cell(row=row_num, column=col, value=label)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = thin
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        # Righe partecipanti
        for ap in parts:
            r = RoomingList.query.get(ap.rooming_list_id) if ap.rooming_list_id else None
            in_rooming = 'Sì' if r else 'No'

            is_cxl     = r.is_cxl if r else False
            is_changed = (r.change_date or (r.latest_changes and str(r.latest_changes).strip())) if r else False
            row_fill   = red_fill if is_cxl else (yellow_fill if is_changed else green_fill)

            for col, (field, _) in enumerate(ALL_ACT_COLS, 1):
                if field == 'diet_restrictions':
                    val = (ap.diet or '') or (r.diet_restrictions if r else '') or ''
                elif field == 'in_rooming':
                    val = in_rooming
                else:
                    val = cell_val(field, r) if r else (ap.last_name if field == 'last_name' else
                                                        ap.first_name if field == 'first_name' else '')
                c = ws.cell(row=row_num, column=col, value=val)
                c.font = norm_font if r else warn_font
                c.fill = row_fill if r else PatternFill('solid', start_color='FFF2CC')
                c.alignment = left; c.border = thin
            ws.row_dimensions[row_num].height = 15
            row_num += 1

    # Larghezze
    for col, (field, _) in enumerate(ALL_ACT_COLS, 1):
        w = COL_WIDTHS.get(field, 14)
        if field == 'diet_restrictions': w = 25
        if field == 'in_rooming': w = 14
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'AllActivities_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rooming_bp.route('/activity-template')
def download_activity_template():
    """Scarica il template Excel per l'import partecipanti attività."""
    import os
    template_path = os.path.join(os.path.dirname(__file__), '..', 'Activity_Template.xlsx')
    return send_file(os.path.abspath(template_path),
                     as_attachment=True,
                     download_name='Activity_Import_Template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rooming_bp.route('/activities')
def activities_page():
    """Pagina gestione attività."""
    from models.models import Activity
    activities = Activity.query.order_by(Activity.date, Activity.name).all()
    # Trova batch più recente per la ricerca partecipanti
    batches = get_batches()
    batch_id = batches[0][0] if batches else ''
    result = [{'id': a.id, 'name': a.name,
               'date': a.date, 'n_participants': len(a.participants)}
              for a in activities]
    return render_template('activities.html', activities=result, batch_id=batch_id)


@rooming_bp.route('/api/activity/<int:activity_id>/participants')
def api_activity_participants(activity_id):
    """Partecipanti di un'attività con dati dalla rooming list."""
    from models.models import Activity, ActivityParticipant
    activity = Activity.query.get_or_404(activity_id)
    result = []
    for ap in sorted(activity.participants,
                     key=lambda x: (x.last_name or '').upper()):
        row = {
            'ap_id':      ap.id,
            'last_name':  ap.last_name or '',
            'first_name': ap.first_name or '',
            'entity':     ap.entity or '',
            'diet':       ap.diet or '',
            'matched':    ap.rooming_list_id is not None,
            'hotel': '', 'arrival_mode': '', 'billing': '',
            'check_in': '', 'check_out': '',
        }
        if ap.rooming_list_id:
            r = RoomingList.query.get(ap.rooming_list_id)
            if r:
                row.update({
                    'hotel':        r.hotel or '',
                    'arrival_mode': r.arrival_mode or '',
                    'billing':      r.billing or '',
                    'check_in':     cell_val('check_in', r),
                    'check_out':    cell_val('check_out', r),
                    'diet':         ap.diet or r.diet_restrictions or '',
                })
        result.append(row)
    return jsonify({
        'activity': activity.name,
        'date': activity.date.strftime('%d/%m/%Y') if activity.date else None,
        'participants': result
    })


@rooming_bp.route('/api/activity-add-participant', methods=['POST'])
def api_activity_add_participant():
    """Aggiunge un partecipante a un'attività."""
    from models.models import Activity, ActivityParticipant
    data         = request.get_json()
    activity_id  = data.get('activity_id')
    rooming_id   = data.get('rooming_list_id')

    activity = Activity.query.get(activity_id)
    if not activity:
        return jsonify({'ok': False, 'error': 'Attività non trovata'})

    # Evita duplicati
    existing = ActivityParticipant.query.filter_by(
        activity_id=activity_id, rooming_list_id=rooming_id).first()
    if existing:
        return jsonify({'ok': False, 'error': 'Partecipante già presente'})

    r = RoomingList.query.get(rooming_id)
    ap = ActivityParticipant(
        activity_id=activity_id,
        rooming_list_id=rooming_id,
        last_name=r.last_name if r else None,
        first_name=r.first_name if r else None,
        diet=r.diet_restrictions if r else None,
    )
    db.session.add(ap)
    db.session.commit()
    return jsonify({'ok': True})


@rooming_bp.route('/api/activity-remove-participant/<int:ap_id>', methods=['DELETE'])
def api_activity_remove_participant(ap_id):
    """Rimuove un partecipante da un'attività."""
    from models.models import ActivityParticipant
    ap = ActivityParticipant.query.get(ap_id)
    if not ap:
        return jsonify({'ok': False, 'error': 'Non trovato'})
    db.session.delete(ap)
    db.session.commit()
    return jsonify({'ok': True})


@rooming_bp.route('/export-activity/<int:activity_id>')
def export_activity(activity_id):
    """Excel lista partecipanti di un'attività."""
    from models.models import Activity
    activity = Activity.query.get_or_404(activity_id)

    COLS = [
        ('last_name',    'Cognome'),
        ('first_name',   'Nome'),
        ('hotel',        'Hotel'),
        ('check_in',     'Check In'),
        ('check_out',    'Check Out'),
        ('arrival_mode', 'Arrivo'),
        ('billing',      'Billing'),
        ('diet',         'Dieta'),
    ]

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = activity.name[:30]

    hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', start_color='1F3864')
    norm_font = Font(name='Calibri', size=10)
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    n_cols = len(COLS)

    # Titolo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1,
                 value=f'{activity.name} — {activity.date.strftime("%d/%m/%Y") if activity.date else ""}')
    tc.font = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(row=2, column=1,
            value=f'Generato il {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  {len(activity.participants)} partecipanti'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    # Header
    for col, (_, label) in enumerate(COLS, 1):
        c = ws.cell(row=3, column=col, value=label)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = thin
    ws.row_dimensions[3].height = 18

    # Righe
    participants = sorted(activity.participants,
                          key=lambda x: (x.last_name or '').upper())
    for rn, ap in enumerate(participants, 4):
        r = RoomingList.query.get(ap.rooming_list_id) if ap.rooming_list_id else None
        vals = {
            'last_name':    ap.last_name or '',
            'first_name':   ap.first_name or '',
            'hotel':        r.hotel if r else '',
            'check_in':     cell_val('check_in', r) if r else '',
            'check_out':    cell_val('check_out', r) if r else '',
            'arrival_mode': r.arrival_mode if r else '',
            'billing':      r.billing if r else '',
            'diet':         ap.diet or (r.diet_restrictions if r else '') or '',
        }
        for col, (field, _) in enumerate(COLS, 1):
            c = ws.cell(row=rn, column=col, value=vals.get(field, ''))
            c.font = norm_font; c.alignment = left; c.border = thin
        ws.row_dimensions[rn].height = 15

    col_widths = {'last_name':20,'first_name':16,'hotel':22,'check_in':12,
                  'check_out':12,'arrival_mode':12,'billing':18,'diet':25}
    for col, (field, _) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(field, 14)
    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    safe = activity.name.replace(' ', '_')[:30]
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'Activity_{safe}_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rooming_bp.route('/api/activities')
def api_activities():
    """Lista tutte le attività."""
    from models.models import Activity
    activities = Activity.query.order_by(Activity.date, Activity.name).all()
    return jsonify([{
        'id':   a.id,
        'name': a.name,
        'date': a.date.strftime('%d/%m/%Y') if a.date else None,
        'n_participants': len(a.participants),
    } for a in activities])

@rooming_bp.route('/api/participant-agenda/<int:rooming_id>')
def api_participant_agenda(rooming_id):
    """Tutte le attività di un partecipante."""
    from models.models import ActivityParticipant
    r = RoomingList.query.get_or_404(rooming_id)
    participations = ActivityParticipant.query.filter_by(rooming_list_id=rooming_id)\
                                              .join(ActivityParticipant.activity)\
                                              .order_by(db.text('activities.date, activities.name'))\
                                              .all()
    return jsonify({
        'name': f'{r.first_name} {r.last_name}',
        'hotel': r.hotel or '',
        'agenda': [{'activity': ap.activity.name,
                    'date': ap.activity.date.strftime('%d/%m/%Y') if ap.activity.date else None}
                   for ap in participations]
    })


@rooming_bp.route('/upload-contracts', methods=['POST'])
def upload_contracts():
    from models.models import HotelContract
    import pandas as pd
    from datetime import date as _date

    f = request.files.get('file')
    if not f or not f.filename.endswith('.xlsx'):
        flash('Seleziona un file .xlsx valido.', 'error')
        return redirect(url_for('rooming.index'))

    df = pd.read_excel(f, sheet_name=0, header=None)

    inserted = 0
    errors   = []
    i = 0
    while i < len(df):
        raw_name = df.iloc[i, 1] if len(df.columns) > 1 else None
        if pd.isna(raw_name):
            i += 1
            continue

        hotel = normalize_hotel(raw_name)
        if not hotel:
            i += 1
            continue

        # Riga date (colonne 2-9) e riga valori (riga successiva)
        if i + 1 >= len(df):
            break

        val_row = df.iloc[i + 1]

        for col_offset, date_str in enumerate(DATES):
            col_idx = col_offset + 2
            try:
                val = val_row.iloc[col_idx] if col_idx < len(val_row) else 0
                rooms = int(val) if pd.notna(val) else 0
            except Exception:
                rooms = 0

            d = _date.fromisoformat(date_str)
            existing = HotelContract.query.filter_by(hotel=hotel, date=d).first()
            if existing:
                existing.rooms = rooms
            else:
                db.session.add(HotelContract(hotel=hotel, date=d, rooms=rooms))
            inserted += 1

        i += 2  # salta la coppia hotel+valori

    # Starling: assicura zero per tutte le date
    for date_str in DATES:
        d = _date.fromisoformat(date_str)
        if not HotelContract.query.filter_by(hotel='Starling', date=d).first():
            db.session.add(HotelContract(hotel='Starling', date=d, rooms=0))

    db.session.commit()
    flash(f'Contratti caricati: {inserted} record aggiornati.', 'success')
    return redirect(url_for('rooming.index'))


@rooming_bp.route('/api/search-participants/<path:batch_id>')
def api_search_participants(batch_id):
    """Cerca partecipanti nel batch per cognome/nome."""
    q = request.args.get('q', '').strip().lower()
    if len(q) < 2:
        return jsonify({'results': []})

    rows = RoomingList.query.filter_by(import_batch=batch_id)\
                            .filter(RoomingList.registration_state != 'CXL')\
                            .all()

    NIGHT_LABELS = ['28/03','29/03','30/03','31/03','01/04','02/04','03/04','04/04']
    NIGHT_FIELDS_S = ['night_sat_28mar','night_sun_29mar','night_mon_30mar','night_tue_31mar',
                      'night_wed_1apr','night_thu_2apr','night_fri_3apr','night_sat_4apr']

    results = []
    for r in rows:
        fn = (r.first_name or '').lower()
        ln = (r.last_name  or '').lower()
        if q in fn or q in ln:
            nights_str = ' '.join(NIGHT_LABELS[i] for i, f in enumerate(NIGHT_FIELDS_S) if getattr(r, f))
            results.append({
                'id':        r.id,
                'last_name': r.last_name or '',
                'first_name':r.first_name or '',
                'hotel':     r.hotel or '',
                'nights':    nights_str or '—',
            })

    results.sort(key=lambda x: (x['last_name'].upper(), x['first_name'].upper()))
    return jsonify({'results': results[:20]})


@rooming_bp.route('/api/associa-partner', methods=['POST'])
def api_associa_partner():
    """Associa manualmente un no-need al suo compagno di stanza e salva per i batch futuri."""
    from models.models import ManualAssociation
    data       = request.get_json()
    row_id     = data.get('row_id')
    partner_id = data.get('partner_id')

    row     = RoomingList.query.get(row_id)
    partner = RoomingList.query.get(partner_id)

    if not row or not partner:
        return jsonify({'ok': False, 'error': 'Record non trovato'})

    NIGHT_FIELDS_A = ['night_sat_28mar','night_sun_29mar','night_mon_30mar','night_tue_31mar',
                      'night_wed_1apr','night_thu_2apr','night_fri_3apr','night_sat_4apr']

    for f in NIGHT_FIELDS_A:
        setattr(row, f, getattr(partner, f))

    if not row.hotel and partner.hotel:
        row.hotel = partner.hotel

    row.room_category = 'shares_room'

    # Salva associazione per batch futuri
    existing = ManualAssociation.query.filter_by(
        last_name=row.last_name, first_name=row.first_name).first()
    if existing:
        existing.partner_last_name  = partner.last_name
        existing.partner_first_name = partner.first_name
    else:
        db.session.add(ManualAssociation(
            last_name=row.last_name,
            first_name=row.first_name,
            partner_last_name=partner.last_name,
            partner_first_name=partner.first_name,
        ))

    db.session.commit()
    return jsonify({'ok': True})



@rooming_bp.route('/api/categories/<path:batch_id>')
def api_categories(batch_id):
    """Ritorna le categorie azienda presenti nel batch, ordinate."""
    cats = db.session.execute(db.text("""
        SELECT DISTINCT company_category
        FROM rooming_list
        WHERE import_batch = :bid
        ORDER BY company_category NULLS LAST
    """), {'bid': batch_id}).fetchall()
    categories = [r[0] for r in cats]
    return jsonify({'categories': categories})



@rooming_bp.route('/report-transfer/<path:batch_id>', methods=['POST'])
def report_transfer(batch_id):
    """Excel transfer: righe=partecipanti, colonne=notti + attività, no CXL.
    Le spouse (status_spouse=Yes) ereditano le notti del titolare via internal_parent_reference.
    """
    from sqlalchemy import or_
    from models.models import Activity, ActivityParticipant
    import re as _re

    hotels = request.form.getlist('hotels')

    q = RoomingList.query.filter_by(import_batch=batch_id)\
                         .filter(RoomingList.registration_state != 'CXL')
    if hotels:
        conditions = [RoomingList.hotel == h for h in hotels]
        q = q.filter(or_(*conditions))

    rows = q.order_by(RoomingList.last_name, RoomingList.first_name).all()

    if not rows:
        flash('Nessun partecipante trovato.', 'error')
        return redirect(url_for('rooming.batch_detail', batch_id=batch_id))

    NIGHTS = [
        ('night_sat_28mar', '28/03'),
        ('night_sun_29mar', '29/03'),
        ('night_mon_30mar', '30/03'),
        ('night_tue_31mar', '31/03'),
        ('night_wed_1apr',  '01/04'),
        ('night_thu_2apr',  '02/04'),
        ('night_fri_3apr',  '03/04'),
        ('night_sat_4apr',  '04/04'),
    ]

    # Mappa internal_reference -> riga per tutto il batch (per lookup spouse)
    all_batch = RoomingList.query.filter_by(import_batch=batch_id)\
                                 .filter(RoomingList.registration_state != 'CXL').all()
    ref_map = {str(r.internal_reference).strip(): r
               for r in all_batch if r.internal_reference}

    # ── Attività: mappa (COGNOME, NOME) -> set di activity_id ──
    # Il match è per nome perché rooming_list_id cambia ad ogni import batch
    activities = Activity.query.order_by(Activity.date, Activity.name).all()
    name_to_activities = {}  # (last_upper, first_upper) -> set(activity_id)
    for act in activities:
        for ap in act.participants:
            key = ((ap.last_name or '').strip().upper(),
                   (ap.first_name or '').strip().upper())
            if key[0]:
                name_to_activities.setdefault(key, set()).add(act.id)

    def get_nights(r):
        """Restituisce dict {field: bool} con le notti effettive del partecipante.
        Se il partecipante non ha notti, cerca il compagno di stanza e usa le sue.
        La logica è simmetrica: non importa chi è il titolare.
        Gestisce anche night_no_need=Yes senza notti proprie."""

        own_nights = {f: bool(getattr(r, f)) for f, _ in NIGHTS}
        if any(own_nights.values()):
            return own_nights

        # Non ha notti — cerca compagno di stanza solo se:
        # - è spouse (status_spouse=Yes), oppure
        # - ha night_no_need=Yes (condivide camera ma non ha notti proprie)
        is_spouse   = (r.status_spouse or '').strip().lower() == 'yes'
        is_no_need  = (r.night_no_need or '').strip().lower() == 'yes'
        if not is_spouse and not is_no_need:
            return own_nights

        # 1. Via internal_parent_reference → cerca titolare
        if r.internal_parent_reference:
            parent = ref_map.get(str(r.internal_parent_reference).strip())
            if parent and any(bool(getattr(parent, f)) for f, _ in NIGHTS):
                return {f: bool(getattr(parent, f)) for f, _ in NIGHTS}

            # 1b. Stesso parent_ref — cerca un altro partecipante con stesso parent che ha notti
            parent_ref_str = str(r.internal_parent_reference).strip()
            for candidate in all_batch:
                if candidate.id == r.id:
                    continue
                if str(candidate.internal_parent_reference or '').strip() == parent_ref_str or \
                   str(candidate.internal_reference or '').strip() == parent_ref_str:
                    cand_nights = {f: bool(getattr(candidate, f)) for f, _ in NIGHTS}
                    if any(cand_nights.values()):
                        return cand_nights

        # 2. Cerca nel campo upgrade del batch chi ha "DOUBLE ROOM WITH / TWIN ROOM WITH <nome o cognome>"
        first = (r.first_name or '').strip().lower()
        last  = (r.last_name or '').strip().lower()
        for candidate in all_batch:
            if candidate.id == r.id:
                continue
            upgrade_text = (candidate.upgrade or '').lower()
            if any(kw in upgrade_text for kw in ('double room with', 'twin room with', 'shared room with')):
                if last in upgrade_text or (first and first in upgrade_text):
                    candidate_nights = {f: bool(getattr(candidate, f)) for f, _ in NIGHTS}
                    if any(candidate_nights.values()):
                        return candidate_nights
                    # Anche il candidato non ha notti — prova il suo parent
                    if candidate.internal_parent_reference:
                        cp = ref_map.get(str(candidate.internal_parent_reference).strip())
                        if cp:
                            cp_nights = {f: bool(getattr(cp, f)) for f, _ in NIGHTS}
                            if any(cp_nights.values()):
                                return cp_nights

        # 3. Cerca nel campo upgrade della spouse stessa
        upgrade_text = (r.upgrade or '').lower()
        if any(kw in upgrade_text for kw in ('double room with', 'twin room with', 'shared room with',
                                              'wife of', 'spouse of')):
            m = _re.search(r'(?:double|twin|shared)\s+room\s+with\s+([\w\s]+)|'
                           r'(?:wife|spouse)\s+of\s+([\w\s]+)',
                           upgrade_text, _re.IGNORECASE)
            if m:
                name_str = (m.group(1) or m.group(2) or '').strip().lower()
                name_parts = name_str.split()
                for candidate in all_batch:
                    if candidate.id == r.id:
                        continue
                    cln = (candidate.last_name or '').lower()
                    cfn = (candidate.first_name or '').lower()
                    if any(p == cln or p == cfn for p in name_parts):
                        candidate_nights = {f: bool(getattr(candidate, f)) for f, _ in NIGHTS}
                        if any(candidate_nights.values()):
                            return candidate_nights

        # 4. Fallback: stesso cognome + stesso hotel + non spouse + ha notti
        spouse_last  = (r.last_name or '').strip().lower()
        spouse_hotel = (r.hotel or '').strip().lower()
        for candidate in all_batch:
            if candidate.id == r.id:
                continue
            cand_last = (candidate.last_name or '').strip().lower()
            cand_hotel = (candidate.hotel or '').strip().lower()
            cand_nights = {f: bool(getattr(candidate, f)) for f, _ in NIGHTS}
            if cand_last == spouse_last and \
               cand_hotel == spouse_hotel and \
               (candidate.status_spouse or '').strip().lower() != 'yes':
                if any(cand_nights.values()):
                    return cand_nights

        # Nessun compagno trovato con notti — vuoto
        return {f: False for f, _ in NIGHTS}

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Transfer'

    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', start_color='1F3864')
    norm_font = Font(name='Calibri', size=10)
    tot_font  = Font(name='Calibri', bold=True, size=10)
    tot_fill  = PatternFill('solid', start_color='D9E1F2')
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    # Colonne attività (etichette corte: "nome (dd/mm)")
    act_labels = []
    for act in activities:
        lbl = act.name or ''
        if act.date:
            lbl += f' ({act.date.strftime("%d/%m")})'
        act_labels.append((act.id, lbl))

    n_cols = 3 + len(NIGHTS) + len(act_labels)
    act_col_start = 4 + len(NIGHTS)   # prima colonna attività (1-based)

    act_fill = PatternFill('solid', start_color='7C3AED')
    act_hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=9)

    # Titolo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value='TRANSFER LIST')
    tc.font = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 26

    # Sottotitolo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    hotel_str = ', '.join(hotels) if hotels else 'tutti gli hotel'
    ts_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.cell(row=2, column=1,
            value=f'Hotel: {hotel_str}  |  {len(rows)} partecipanti  |  {ts_str}'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    # Intestazioni
    headers = ['Cognome', 'Nome', 'Hotel'] + [label for _, label in NIGHTS]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = thin
    # Intestazioni attività
    for ai, (act_id, lbl) in enumerate(act_labels):
        c = ws.cell(row=3, column=act_col_start + ai, value=lbl)
        c.font = act_hdr_font; c.fill = act_fill
        c.alignment = center; c.border = thin
    ws.row_dimensions[3].height = 20

    # Righe partecipanti + contatori totale
    totals_nights = [0] * len(NIGHTS)
    totals_acts   = [0] * len(act_labels)
    for rn, r in enumerate(rows, 4):
        ws.cell(row=rn, column=1, value=r.last_name or '').font = norm_font
        ws.cell(row=rn, column=1).border = thin
        ws.cell(row=rn, column=1).alignment = left
        ws.cell(row=rn, column=2, value=r.first_name or '').font = norm_font
        ws.cell(row=rn, column=2).border = thin
        ws.cell(row=rn, column=2).alignment = left
        ws.cell(row=rn, column=3, value=r.hotel or '').font = norm_font
        ws.cell(row=rn, column=3).border = thin
        ws.cell(row=rn, column=3).alignment = left

        for ni, (field, _) in enumerate(NIGHTS):
            # Eredita notti del parent solo se night_no_need=Yes + parent_ref valorizzato
            is_no_need = (r.night_no_need or '').strip().lower() == 'yes'
            has_parent = bool(r.internal_parent_reference)
            if is_no_need and has_parent and not getattr(r, field):
                parent = ref_map.get(str(r.internal_parent_reference).strip())
                night_val = bool(getattr(parent, field)) if parent else False
            else:
                night_val = bool(getattr(r, field))
            val = 'X' if night_val else ''
            c = ws.cell(row=rn, column=4+ni, value=val)
            c.font = norm_font; c.border = thin; c.alignment = center
            if val:
                totals_nights[ni] += 1

        # Colonne attività (match per nome)
        name_key = ((r.last_name or '').strip().upper(),
                     (r.first_name or '').strip().upper())
        part_acts = name_to_activities.get(name_key, set())
        for ai, (act_id, _) in enumerate(act_labels):
            val = 'X' if act_id in part_acts else ''
            c = ws.cell(row=rn, column=act_col_start + ai, value=val)
            c.font = norm_font; c.border = thin; c.alignment = center
            if val:
                totals_acts[ai] += 1

        ws.row_dimensions[rn].height = 15

    # Riga totale
    tot_row = len(rows) + 4
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=3)
    tc = ws.cell(row=tot_row, column=1, value='TOTALE')
    tc.font = tot_font; tc.fill = tot_fill; tc.alignment = left; tc.border = thin
    for ni, t in enumerate(totals_nights):
        c = ws.cell(row=tot_row, column=4+ni, value=t)
        c.font = tot_font; c.fill = tot_fill
        c.alignment = center; c.border = thin
    for ai, t in enumerate(totals_acts):
        c = ws.cell(row=tot_row, column=act_col_start + ai, value=t)
        c.font = tot_font; c.fill = tot_fill
        c.alignment = center; c.border = thin
    ws.row_dimensions[tot_row].height = 18

    # Larghezze colonne
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    for ni in range(len(NIGHTS)):
        ws.column_dimensions[get_column_letter(4+ni)].width = 8
    for ai in range(len(act_labels)):
        ws.column_dimensions[get_column_letter(act_col_start + ai)].width = 18

    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    hotel_tag = hotels[0].replace(' ', '_')[:20] if len(hotels)==1 else 'multi'
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'Transfer_{hotel_tag}_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rooming_bp.route('/api/billings/<path:batch_id>')
def api_billings(batch_id):
    """Ritorna i valori billing presenti nel batch, ordinati."""
    rows = db.session.execute(db.text("""
        SELECT DISTINCT billing
        FROM rooming_list
        WHERE import_batch = :bid
        ORDER BY billing NULLS LAST
    """), {'bid': batch_id}).fetchall()
    return jsonify({'billings': [r[0] for r in rows]})


@rooming_bp.route('/report-billing/<path:batch_id>', methods=['POST'])
def report_billing(batch_id):
    """Excel con partecipanti filtrati per billing, raggruppati per hotel."""
    from sqlalchemy import or_
    billings = request.form.getlist('billings')
    if not billings:
        flash('Seleziona almeno un valore billing.', 'error')
        return redirect(url_for('rooming.batch_detail', batch_id=batch_id))

    conditions = []
    for b in billings:
        if b == '__blank__':
            conditions.append(RoomingList.billing == None)
        else:
            conditions.append(RoomingList.billing == b)

    rows = RoomingList.query.filter_by(import_batch=batch_id)\
                            .filter(or_(*conditions))\
                            .order_by(RoomingList.hotel, RoomingList.last_name,
                                      RoomingList.first_name).all()

    if not rows:
        flash('Nessun partecipante trovato per i valori billing selezionati.', 'error')
        return redirect(url_for('rooming.batch_detail', batch_id=batch_id))

    def _sort(r):
        return (r.hotel or '', 1 if r.is_cxl else 0,
                str(r.last_name or '').upper(), str(r.first_name or '').upper())
    rows = sorted(rows, key=_sort)

    bill_labels = [b if b != '__blank__' else '(blank)' for b in billings]
    title_str   = ' · '.join(bill_labels)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Report Billing'

    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', start_color='1F3864')
    norm_font = Font(name='Calibri', size=10)
    green_fill  = PatternFill('solid', start_color='C6EFCE')
    yellow_fill = PatternFill('solid', start_color='FFEB9C')
    red_fill    = PatternFill('solid', start_color='FFC7CE')
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    n_cols = len(PREVIEW_COLS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value=f'REPORT BILLING — {title_str.upper()}')
    tc.font = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ts_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.cell(row=2, column=1,
            value=f'Generato il {ts_str}  |  {len(rows)} partecipanti'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    current_hotel = None
    row_num = 3

    for r in rows:
        hotel = r.hotel or 'SENZA HOTEL'
        if hotel != current_hotel:
            current_hotel = hotel
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=n_cols)
            hc = ws.cell(row=row_num, column=1, value=hotel.upper())
            hc.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
            hc.fill = PatternFill('solid', start_color='2F5496')
            hc.alignment = left
            ws.row_dimensions[row_num].height = 20
            row_num += 1
            for col, (field, label) in enumerate(PREVIEW_COLS, 1):
                c = ws.cell(row=row_num, column=col, value=label)
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = center; c.border = thin
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        is_changed = r.change_date or (r.latest_changes and str(r.latest_changes).strip())
        row_fill = red_fill if r.is_cxl else (yellow_fill if is_changed else green_fill)

        for col, (field, _) in enumerate(PREVIEW_COLS, 1):
            c = ws.cell(row=row_num, column=col, value=cell_val(field, r))
            c.font = norm_font; c.fill = row_fill
            c.alignment = left; c.border = thin
        ws.row_dimensions[row_num].height = 15
        row_num += 1

    for col, (field, _) in enumerate(PREVIEW_COLS, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(field, 14)
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = title_str.replace(' ', '_').replace('/', '-')[:30]
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'Report_Billing_{safe}_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rooming_bp.route('/api/arrivals/<path:batch_id>')
def api_arrivals(batch_id):
    rows = db.session.execute(db.text("""
        SELECT DISTINCT arrival_mode
        FROM rooming_list
        WHERE import_batch = :bid
        ORDER BY arrival_mode NULLS LAST
    """), {'bid': batch_id}).fetchall()
    return jsonify({'arrivals': [r[0] for r in rows]})


@rooming_bp.route('/report-arrival/<path:batch_id>', methods=['POST'])
def report_arrival(batch_id):
    from sqlalchemy import or_
    arrivals = request.form.getlist('arrivals')
    if not arrivals:
        flash('Seleziona almeno un valore.', 'error')
        return redirect(url_for('rooming.batch_detail', batch_id=batch_id))

    conditions = [RoomingList.arrival_mode == None if a == '__blank__'
                  else RoomingList.arrival_mode == a for a in arrivals]

    rows = RoomingList.query.filter_by(import_batch=batch_id)\
                            .filter(or_(*conditions))\
                            .order_by(RoomingList.hotel, RoomingList.last_name,
                                      RoomingList.first_name).all()

    if not rows:
        flash('Nessun partecipante trovato.', 'error')
        return redirect(url_for('rooming.batch_detail', batch_id=batch_id))

    def _sort(r):
        return (r.hotel or '', 1 if r.is_cxl else 0,
                str(r.last_name or '').upper(), str(r.first_name or '').upper())
    rows = sorted(rows, key=_sort)

    labels    = [a if a != '__blank__' else '(blank)' for a in arrivals]
    title_str = ' · '.join(labels)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Report Arrival'

    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', start_color='1F3864')
    norm_font = Font(name='Calibri', size=10)
    green_fill  = PatternFill('solid', start_color='C6EFCE')
    yellow_fill = PatternFill('solid', start_color='FFEB9C')
    red_fill    = PatternFill('solid', start_color='FFC7CE')
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    n_cols = len(PREVIEW_COLS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value=f'REPORT ARRIVAL — {title_str.upper()}')
    tc.font = Font(name='Calibri', bold=True, size=14, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ts_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.cell(row=2, column=1,
            value=f'Generato il {ts_str}  |  {len(rows)} partecipanti'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    current_hotel = None
    row_num = 3

    for r in rows:
        hotel = r.hotel or 'SENZA HOTEL'
        if hotel != current_hotel:
            current_hotel = hotel
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=n_cols)
            hc = ws.cell(row=row_num, column=1, value=hotel.upper())
            hc.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
            hc.fill = PatternFill('solid', start_color='2F5496')
            hc.alignment = left
            ws.row_dimensions[row_num].height = 20
            row_num += 1
            for col, (field, label) in enumerate(PREVIEW_COLS, 1):
                c = ws.cell(row=row_num, column=col, value=label)
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = center; c.border = thin
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        is_changed = r.change_date or (r.latest_changes and str(r.latest_changes).strip())
        row_fill = red_fill if r.is_cxl else (yellow_fill if is_changed else green_fill)

        for col, (field, _) in enumerate(PREVIEW_COLS, 1):
            c = ws.cell(row=row_num, column=col, value=cell_val(field, r))
            c.font = norm_font; c.fill = row_fill
            c.alignment = left; c.border = thin
        ws.row_dimensions[row_num].height = 15
        row_num += 1

    for col, (field, _) in enumerate(PREVIEW_COLS, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(field, 14)
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = title_str.replace(' ', '_').replace('/', '-')[:30]
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'Report_Arrival_{safe}_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rooming_bp.route('/api/occupancy/<path:batch_id>')
def api_occupancy(batch_id):
    """Ritorna JSON con CONFIRMED vs BY CONTRACT per hotel x data."""
    from models.models import HotelContract
    from datetime import date as _date

    rows = RoomingList.query.filter_by(import_batch=batch_id)\
                            .filter(RoomingList.registration_state != 'CXL')\
                            .all()

    # Notti field → date
    NIGHT_TO_DATE = {
        'night_sat_28mar': '2026-03-28',
        'night_sun_29mar': '2026-03-29',
        'night_mon_30mar': '2026-03-30',
        'night_tue_31mar': '2026-03-31',
        'night_wed_1apr':  '2026-04-01',
        'night_thu_2apr':  '2026-04-02',
        'night_fri_3apr':  '2026-04-03',
        'night_sat_4apr':  '2026-04-04',
    }

    hotels = sorted(set(r.hotel for r in rows if r.hotel))

    # Costruisce confirmed[hotel][date] = count
    confirmed = {h: {d: 0 for d in DATES} for h in hotels}
    for r in rows:
        if not r.hotel:
            continue
        for field, date_str in NIGHT_TO_DATE.items():
            if getattr(r, field):
                confirmed[r.hotel][date_str] += 1

    # Carica contratti
    contracts = HotelContract.query.all()
    contract_map = {(c.hotel, str(c.date)): c.rooms for c in contracts}

    result = {}
    for h in hotels:
        result[h] = {
            'confirmed': [confirmed[h].get(d, 0) for d in DATES],
            'contract':  [contract_map.get((h, d), 0) for d in DATES],
        }

    return jsonify({
        'hotels':       hotels,
        'dates':        DATE_LABELS,
        'data':         result,
        'has_contracts': len(contracts) > 0,
    })


# ── Original File ─────────────────────────────────────────────────────────────

# Colonne "file originale": tutte quelle del COL_MAP + campi derivati
ORIGINAL_COLS = [
    # Riga 1
    ('title',                     'Title'),
    ('last_name',                 'Last Name'),
    ('first_name',                'First Name'),
    ('comment',                   'Comment'),
    ('email',                     'Email'),
    ('phone',                     'Phone'),
    ('billing',                   'Billing'),
    ('upgrade',                   'Upgrade'),
    ('check_in',                  'Check In'),
    ('check_out',                 'Check Out'),
    ('company_name',              'Company'),
    ('company_country',           'Country'),
    ('job_position',              'Job Position'),
    ('company_category',          'Category'),
    ('company_subcategory',       'Sub-Category'),
    # Riga 2
    ('registration_state',        'Reg. State'),
    ('hotel',                     'Hotel'),
    ('nexus_bd',                  'NEXUS BD'),
    ('night_no_need',             'No need'),
    ('night_sat_28mar',           'Sat 28/03'),
    ('night_sun_29mar',           'Sun 29/03'),
    ('night_mon_30mar',           'Mon 30/03'),
    ('night_tue_31mar',           'Tue 31/03'),
    ('night_wed_1apr',            'Wed 01/04'),
    ('night_thu_2apr',            'Thu 02/04'),
    ('night_fri_3apr',            'Fri 03/04'),
    ('night_sat_4apr',            'Sat 04/04'),
    ('arrival_mode',              'Arrival'),
    ('need_smooth_checkin',       'Smooth CI'),
    ('diet_restrictions',         'Diet'),
    # Riga 3
    ('need_visa',                 'Visa?'),
    ('visa_birth_date',           'Birth Date'),
    ('visa_birth_place',          'Birth Place'),
    ('visa_passport',             'Passport N.'),
    ('visa_delivery_date',        'Doc Issued'),
    ('visa_expiration_date',      'Doc Expiry'),
    ('visa_company_address',      'Company Address'),
    ('comment',                   'Comment'),
    ('internal_reference',        'Internal Ref'),
    ('participant_number',        'Participant N.'),
    ('ean8_barcode',              'EAN8'),
    ('status_vp_bd',              'VP/BD'),
    ('status_organisator',        'Organisator'),
    ('status_board_nai',          'Board NAI'),
    ('status_climate_day',        'Climate Day'),
    # Riga 4
    ('status_spouse',             'Spouse Flag'),
    ('spouse_name',               'Spouse Name'),
    ('is_parent_manager',         'Parent Mgr'),
    ('registered_colleagues',     'Colleagues'),
    ('latest_changes',            'Latest Changes'),
    ('change_type',               'Change Type'),
    ('change_date',               'Change Date'),
    ('file_timestamp',            'File Timestamp'),
]

# Campi su cui permettere il filtro query
ORIGINAL_FILTER_FIELDS = [
    ('last_name',           'Last Name'),
    ('first_name',          'First Name'),
    ('hotel',               'Hotel'),
    ('registration_state',  'Stato'),
    ('company_name',        'Company'),
    ('company_category',    'Category'),
    ('billing',             'Billing'),
    ('arrival_mode',        'Arrival'),
    ('change_type',         'Change Type'),
    ('status_spouse',       'Spouse'),
    ('need_visa',           'Visa?'),
    ('night_sat_28mar',     '28/03'),
    ('night_sun_29mar',     '29/03'),
    ('night_mon_30mar',     '30/03'),
    ('night_tue_31mar',     '31/03'),
    ('night_wed_1apr',      '01/04'),
    ('night_thu_2apr',      '02/04'),
    ('night_fri_3apr',      '03/04'),
    ('night_sat_4apr',      '04/04'),
]


@rooming_bp.route('/original-file')
def original_file():
    """Pagina con preview completa del rooming list originale (ultimo batch)."""
    batches = get_batches()
    if not batches:
        flash('Nessun batch disponibile.', 'error')
        return redirect(url_for('rooming.index'))
    batch_id = batches[0][0]
    batch_ts = batches[0][2]
    total    = batches[0][1]
    return render_template('original_file.html',
                           batch_id=batch_id,
                           batch_ts=batch_ts,
                           total=total,
                           cols=ORIGINAL_COLS,
                           filter_fields=ORIGINAL_FILTER_FIELDS)


@rooming_bp.route('/api/original-file/<path:batch_id>', methods=['POST'])
def api_original_file(batch_id):
    """JSON paginato con filtri per la pagina file originale."""
    from sqlalchemy import or_, and_
    data     = request.get_json()
    page     = int(data.get('page', 1))
    per_page = int(data.get('per_page', 100))
    filters  = data.get('filters', [])   # [{field, op, val}, ...]
    logic    = data.get('logic', 'AND')
    cols     = data.get('cols', [f for f, _ in ORIGINAL_COLS])

    q = RoomingList.query.filter_by(import_batch=batch_id)

    conds = []
    for f in filters:
        field = f.get('field', '')
        op    = f.get('op', 'contains')
        val   = f.get('val', '')
        col   = getattr(RoomingList, field, None)
        if col is None:
            continue
        if op == 'eq':          conds.append(col == val)
        elif op == 'neq':       conds.append(col != val)
        elif op == 'contains':  conds.append(col.ilike(f'%{val}%'))
        elif op == 'not_empty': conds.append(col != None)
        elif op == 'empty':     conds.append(col == None)
    if conds:
        q = q.filter(and_(*conds) if logic == 'AND' else or_(*conds))

    total  = q.count()
    rows   = q.order_by(RoomingList.hotel, RoomingList.last_name,
                        RoomingList.first_name)\
              .offset((page - 1) * per_page)\
              .limit(per_page)\
              .all()

    col_labels = {f: l for f, l in ORIGINAL_COLS}
    result = []
    for r in rows:
        row_data = {'_is_cxl': r.is_cxl,
                    '_has_changes': bool(r.change_date or
                                         (r.latest_changes and str(r.latest_changes).strip())),
                    'internal_reference': r.internal_reference or ''}
        for f in cols:
            row_data[f] = cell_val(f, r)
        result.append(row_data)

    return jsonify({
        'rows':   result,
        'total':  total,
        'page':   page,
        'pages':  (total + per_page - 1) // per_page,
        'labels': col_labels,
    })


@rooming_bp.route('/api/export-original/<path:batch_id>', methods=['POST'])
def api_export_original(batch_id):
    """Esporta in Excel le righe filtrate del file originale."""
    from sqlalchemy import or_, and_
    data    = request.get_json()
    filters = data.get('filters', [])
    logic   = data.get('logic', 'AND')

    q = RoomingList.query.filter_by(import_batch=batch_id)
    conds = []
    for f in filters:
        field = f.get('field', '')
        op    = f.get('op', 'contains')
        val   = f.get('val', '')
        col   = getattr(RoomingList, field, None)
        if col is None:
            continue
        if op == 'eq':          conds.append(col == val)
        elif op == 'neq':       conds.append(col != val)
        elif op == 'contains':  conds.append(col.ilike(f'%{val}%'))
        elif op == 'not_empty': conds.append(col != None)
        elif op == 'empty':     conds.append(col == None)
    if conds:
        q = q.filter(and_(*conds) if logic == 'AND' else or_(*conds))

    rows = q.order_by(RoomingList.hotel, RoomingList.last_name,
                      RoomingList.first_name).all()

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'File Originale'

    hdr_font    = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
    hdr_fill    = PatternFill('solid', start_color='1F3864')
    norm_font   = Font(name='Calibri', size=9)
    green_fill  = PatternFill('solid', start_color='C6EFCE')
    yellow_fill = PatternFill('solid', start_color='FFEB9C')
    red_fill    = PatternFill('solid', start_color='FFC7CE')
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    n_cols = len(ORIGINAL_COLS)

    # Titolo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value='FILE ORIGINALE — ROOMING LIST')
    tc.font = Font(name='Calibri', bold=True, size=13, color='1F3864')
    tc.alignment = center
    ws.row_dimensions[1].height = 24

    # Sottotitolo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ts_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    filter_desc = f'{len(filters)} filtri attivi' if filters else 'nessun filtro'
    ws.cell(row=2, column=1,
            value=f'Generato il {ts_str}  |  {len(rows)} righe  |  {filter_desc}'
            ).font = Font(name='Calibri', size=9, italic=True)
    ws.row_dimensions[2].height = 14

    # Header
    for col, (field, label) in enumerate(ORIGINAL_COLS, 1):
        c = ws.cell(row=3, column=col, value=label)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = thin
    ws.row_dimensions[3].height = 20

    for rn, r in enumerate(rows, 4):
        is_cxl     = r.is_cxl
        is_changed = r.change_date or (r.latest_changes and str(r.latest_changes).strip())
        row_fill   = red_fill if is_cxl else (yellow_fill if is_changed else green_fill)
        for col, (field, _) in enumerate(ORIGINAL_COLS, 1):
            c = ws.cell(row=rn, column=col, value=cell_val(field, r))
            c.font = norm_font; c.fill = row_fill
            c.alignment = left; c.border = thin
        ws.row_dimensions[rn].height = 14

    # Larghezze
    night_fields = {'night_sat_28mar','night_sun_29mar','night_mon_30mar','night_tue_31mar',
                    'night_wed_1apr','night_thu_2apr','night_fri_3apr','night_sat_4apr',
                    'night_no_need','night_other'}
    for col, (field, _) in enumerate(ORIGINAL_COLS, 1):
        if field in night_fields:
            w = 7
        else:
            w = COL_WIDTHS.get(field, 14)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True,
                     download_name=f'FileOriginale_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Override manuali ──────────────────────────────────────────────────────────

# Campi editabili con etichetta
EDITABLE_FIELDS = [
    ('title',                     'Title'),
    ('last_name',                 'Last Name'),
    ('first_name',                'First Name'),
    ('comment',                   'Comment'),
    ('email',                     'Email'),
    ('phone',                     'Phone'),
    ('billing',                   'Billing'),
    ('upgrade',                   'Upgrade'),
    ('company_name',              'Company'),
    ('company_country',           'Country'),
    ('job_position',              'Job Position'),
    ('company_category',          'Category'),
    ('company_subcategory',       'Sub-Category'),
    ('registration_state',        'Reg. State'),
    ('hotel',                     'Hotel'),
    ('nexus_bd',                  'NEXUS BD'),
    ('night_no_need',             'No need'),
    ('night_sat_28mar',           'Notte 28/03'),
    ('night_sun_29mar',           'Notte 29/03'),
    ('night_mon_30mar',           'Notte 30/03'),
    ('night_tue_31mar',           'Notte 31/03'),
    ('night_wed_1apr',            'Notte 01/04'),
    ('night_thu_2apr',            'Notte 02/04'),
    ('night_fri_3apr',            'Notte 03/04'),
    ('night_sat_4apr',            'Notte 04/04'),
    ('arrival_mode',              'Arrival'),
    ('need_smooth_checkin',       'Smooth CI'),
    ('diet_restrictions',         'Diet'),
    ('need_visa',                 'Visa?'),
    ('visa_birth_date',           'Birth Date'),
    ('visa_birth_place',          'Birth Place'),
    ('visa_passport',             'Passport N.'),
    ('visa_delivery_date',        'Doc Issued'),
    ('visa_expiration_date',      'Doc Expiry'),
    ('visa_company_address',      'Company Address'),
    ('status_vp_bd',              'VP/BD'),
    ('status_organisator',        'Organisator'),
    ('status_board_nai',          'Board NAI'),
    ('status_climate_day',        'Climate Day'),
    ('status_spouse',             'Spouse'),
    ('is_parent_manager',         'Parent Mgr'),
    ('registered_colleagues',     'Colleagues'),
    ('latest_changes',            'Latest Changes'),
]


@rooming_bp.route('/api/override', methods=['POST'])
def api_override():
    """Salva o aggiorna un override manuale."""
    from flask import session as flask_session
    from models.models import ManualOverride
    data     = request.get_json()
    ref      = (data.get('internal_reference') or '').strip()
    field    = data.get('field', '').strip()
    new_val  = data.get('value', '')
    username = flask_session.get('username', 'unknown')

    if not ref or not field:
        return jsonify({'ok': False, 'error': 'Dati mancanti'})

    # Leggi valore attuale dal batch più recente
    batches = get_batches()
    original_val = ''
    if batches:
        row = RoomingList.query.filter_by(
            import_batch=batches[0][0], internal_reference=ref).first()
        if row:
            original_val = str(getattr(row, field) or '')

    # Upsert override
    ov = ManualOverride.query.filter_by(
        internal_reference=ref, field=field).first()
    if ov:
        ov.original_value = original_val
        ov.override_value = new_val
        ov.modified_at    = datetime.now()
        ov.modified_by    = username
    else:
        ov = ManualOverride(
            internal_reference=ref,
            field=field,
            original_value=original_val,
            override_value=new_val,
            modified_at=datetime.now(),
            modified_by=username,
        )
        db.session.add(ov)

    # Applica subito al batch corrente + aggiorna latest_changes
    if batches:
        row = RoomingList.query.filter_by(
            import_batch=batches[0][0], internal_reference=ref).first()
        if row:
            setattr(row, field, new_val)
            # Marca il record come modificato manualmente -> diventa giallo
            today_str = datetime.now().strftime('%d/%m/%Y')
            row.latest_changes = f'MANUAL EDIT - {today_str} ({username})'
            row.change_type    = 'MANUAL EDIT'
            row.change_date    = datetime.now().date()

    db.session.commit()
    return jsonify({'ok': True})


@rooming_bp.route('/api/override-delete', methods=['POST'])
def api_override_delete():
    """Ripristina un campo al valore originale del batch."""
    from models.models import ManualOverride
    data  = request.get_json()
    ref   = (data.get('internal_reference') or '').strip()
    field = data.get('field', '').strip()

    ov = ManualOverride.query.filter_by(
        internal_reference=ref, field=field).first()
    if not ov:
        return jsonify({'ok': False, 'error': 'Override non trovato'})

    # Ripristina il valore originale nel batch corrente
    batches = get_batches()
    if batches:
        row = RoomingList.query.filter_by(
            import_batch=batches[0][0], internal_reference=ref).first()
        if row:
            setattr(row, field, ov.original_value)

    db.session.delete(ov)
    db.session.commit()
    return jsonify({'ok': True})


@rooming_bp.route('/api/overrides/<internal_reference>')
def api_get_overrides(internal_reference):
    """Restituisce tutti gli override attivi per un partecipante."""
    from models.models import ManualOverride
    ovs = ManualOverride.query.filter_by(
        internal_reference=internal_reference).all()
    return jsonify({
        'overrides': {ov.field: {
            'value':      ov.override_value,
            'original':   ov.original_value,
            'modified_at': ov.modified_at.strftime('%d/%m/%Y %H:%M'),
            'modified_by': ov.modified_by,
        } for ov in ovs}
    })


@rooming_bp.route('/api/participant/<internal_reference>')
def api_get_participant(internal_reference):
    """Restituisce i dati attuali di un partecipante (con override applicati)."""
    batches = get_batches()
    if not batches:
        return jsonify({'ok': False, 'error': 'Nessun batch'})
    row = RoomingList.query.filter_by(
        import_batch=batches[0][0],
        internal_reference=internal_reference).first()
    if not row:
        return jsonify({'ok': False, 'error': 'Partecipante non trovato'})

    from models.models import ManualOverride
    ovs = {ov.field: ov.override_value
           for ov in ManualOverride.query.filter_by(
               internal_reference=internal_reference).all()}

    # Campi con selezione da tendina
    SELECT_FIELDS = {'hotel', 'billing', 'company_category', 'company_subcategory'}

    # Carica opzioni distinte per i campi select
    field_options = {}
    for f in SELECT_FIELDS:
        col = getattr(RoomingList, f, None)
        if col is not None:
            vals = db.session.execute(db.text(
                f"SELECT DISTINCT {f} FROM rooming_list WHERE import_batch = :bid AND {f} IS NOT NULL ORDER BY {f}"
            ), {'bid': batches[0][0]}).fetchall()
            field_options[f] = [r[0] for r in vals if r[0]]

    fields_dict = {}
    for field, label in EDITABLE_FIELDS:
        entry = {
            'label':    label,
            'value':    cell_val(field, row),
            'override': field in ovs,
            'type':     'select' if field in SELECT_FIELDS else 'text',
        }
        if field in SELECT_FIELDS:
            entry['options'] = field_options.get(field, [])
        fields_dict[field] = entry

    return jsonify({
        'ok':        True,
        'name':      f'{row.first_name or ""} {row.last_name or ""}'.strip(),
        'ref':       internal_reference,
        'fields':    fields_dict,
    })


@rooming_bp.route('/overrides-log')
def overrides_log():
    """Pagina log di tutti gli override."""
    from models.models import ManualOverride
    ovs = ManualOverride.query.order_by(
        ManualOverride.modified_at.desc()).all()

    # Arricchisci con nome partecipante dall'ultimo batch
    batches = get_batches()
    ref_to_name = {}
    if batches:
        rows = RoomingList.query.filter_by(import_batch=batches[0][0]).all()
        ref_to_name = {
            str(r.internal_reference): f'{r.first_name or ""} {r.last_name or ""}'.strip()
            for r in rows if r.internal_reference
        }

    field_labels = {f: l for f, l in EDITABLE_FIELDS}

    log_entries = [{
        'ref':          ov.internal_reference,
        'name':         ref_to_name.get(str(ov.internal_reference), '—'),
        'field':        field_labels.get(ov.field, ov.field),
        'original':     ov.original_value or '—',
        'override':     ov.override_value or '—',
        'modified_at':  ov.modified_at.strftime('%d/%m/%Y %H:%M'),
        'modified_by':  ov.modified_by,
    } for ov in ovs]

    return render_template('overrides_log.html', log_entries=log_entries,
                           total=len(log_entries))


# ── Aggiunta manuale partecipante ─────────────────────────────────────────────

@rooming_bp.route('/api/add-participant', methods=['POST'])
def api_add_participant():
    """Aggiunge un partecipante manualmente al batch corrente."""
    from flask import session as flask_session
    data     = request.get_json()
    username = flask_session.get('username', 'unknown')

    batches = get_batches()
    if not batches:
        return jsonify({'ok': False, 'error': 'Nessun batch disponibile'})

    batch_id = batches[0][0]

    # Genera internal_reference fittizio univoco
    existing = db.session.execute(db.text(
        "SELECT internal_reference FROM rooming_list "
        "WHERE internal_reference LIKE 'MANUAL-%' "
        "ORDER BY internal_reference DESC LIMIT 1"
    )).fetchone()

    if existing:
        try:
            last_num = int(existing[0].replace('MANUAL-', ''))
            new_ref  = f'MANUAL-{last_num + 1:03d}'
        except Exception:
            new_ref = 'MANUAL-001'
    else:
        new_ref = 'MANUAL-001'

    today = datetime.now().date()
    today_str = datetime.now().strftime('%d/%m/%Y')

    row = RoomingList(
        import_batch       = batch_id,
        internal_reference = new_ref,
        change_type        = 'MANUAL ADD',
        change_date        = today,
        latest_changes     = f'MANUAL ADD - {today_str} ({username})',
        registration_state = 'OK',
    )

    # Applica i campi passati dal form
    ALLOWED = {f for f, _ in EDITABLE_FIELDS}
    for field, value in data.get('fields', {}).items():
        if field in ALLOWED and hasattr(row, field):
            setattr(row, field, value if value != '' else None)

    db.session.add(row)
    db.session.commit()

    return jsonify({'ok': True, 'ref': new_ref})


@rooming_bp.route('/api/create-activity', methods=['POST'])
def api_create_activity():
    """Crea una nuova attività manualmente."""
    from models.models import Activity
    data = request.get_json()
    name = (data.get('name') or '').strip()
    date_str = (data.get('date') or '').strip()

    if not name:
        return jsonify({'ok': False, 'error': 'Nome obbligatorio'})

    activity_date = None
    if date_str:
        try:
            from datetime import datetime as _dt
            activity_date = _dt.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            pass

    # Controlla duplicati
    existing = Activity.query.filter_by(name=name, date=activity_date).first()
    if existing:
        return jsonify({'ok': False, 'error': 'Attività già esistente con stesso nome e data'})

    activity = Activity(name=name, date=activity_date)
    db.session.add(activity)
    db.session.commit()
    return jsonify({'ok': True, 'id': activity.id})


# ── AI ASSISTANT (Claude) ────────────────────────────────────────────────────

AI_NIGHT_MAP = {
    'night_sat_28mar': ['28/3', '28 marzo', '28mar', 'sabato 28'],
    'night_sun_29mar': ['29/3', '29 marzo', '29mar', 'domenica 29'],
    'night_mon_30mar': ['30/3', '30 marzo', '30mar', 'lunedì 30', 'lunedi 30'],
    'night_tue_31mar': ['31/3', '31 marzo', '31mar', 'martedì 31', 'martedi 31'],
    'night_wed_1apr':  ['1/4', '1 aprile', '1apr', 'mercoledì 1', 'mercoledi 1'],
    'night_thu_2apr':  ['2/4', '2 aprile', '2apr', 'giovedì 2', 'giovedi 2'],
    'night_fri_3apr':  ['3/4', '3 aprile', '3apr', 'venerdì 3', 'venerdi 3'],
    'night_sat_4apr':  ['4/4', '4 aprile', '4apr', 'sabato 4'],
}

AI_NIGHT_LABELS = {
    'night_sat_28mar': '28/3', 'night_sun_29mar': '29/3',
    'night_mon_30mar': '30/3', 'night_tue_31mar': '31/3',
    'night_wed_1apr': '1/4', 'night_thu_2apr': '2/4',
    'night_fri_3apr': '3/4', 'night_sat_4apr': '4/4',
}


def _ai_build_summary(rows):
    """Costruisce un riassunto statistico del batch."""
    from collections import Counter
    active = [r for r in rows if not r.is_cxl]
    cxl = [r for r in rows if r.is_cxl]

    hotel_counts = Counter(r.hotel or '(vuoto)' for r in active)
    hotel_lines = ', '.join(f"{h}:{c}" for h, c in hotel_counts.most_common())

    night_counts = {}
    for fld, lbl in AI_NIGHT_LABELS.items():
        night_counts[lbl] = sum(1 for r in active
                                if (getattr(r, fld) or '').strip().upper() == 'YES')
    night_lines = ', '.join(f"{lbl}:{c}" for lbl, c in night_counts.items())

    country_counts = Counter(r.company_country or '?' for r in active)
    country_lines = ', '.join(f"{c}:{n}" for c, n in country_counts.most_common(10))

    diets = [r for r in active if r.diet_restrictions and r.diet_restrictions.strip()]

    return (f"RIEPILOGO: {len(active)} attivi, {len(cxl)} CXL\n"
            f"Hotel: {hotel_lines}\n"
            f"Notti: {night_lines}\n"
            f"Paesi top: {country_lines}\n"
            f"Con dieta: {len(diets)}\n"
            f"Lista hotel: {', '.join(h for h, _ in hotel_counts.most_common())}")


def _ai_detect_activity(question):
    """Rileva se la domanda riguarda un'attività. Restituisce (Activity, partecipanti_csv) o (None, None)."""
    from models.models import Activity, ActivityParticipant
    q_lower = question.lower()

    # Keyword che indicano domanda su attività
    activity_keywords = ['attività', 'attivita', 'activity', 'partecipa', 'iscritto',
                         'iscritti', 'evento', 'sessione', 'workshop', 'meeting',
                         'cena', 'pranzo', 'dinner', 'lunch', 'gala', 'tour',
                         'escursione', 'visita']
    if not any(kw in q_lower for kw in activity_keywords):
        return None, None

    # Cerca quale attività matcha
    activities = Activity.query.all()
    matched_activity = None
    best_score = 0
    for act in activities:
        act_words = [w.lower() for w in (act.name or '').split() if len(w) > 2]
        score = sum(1 for w in act_words if w in q_lower)
        if score > best_score:
            best_score = score
            matched_activity = act

    if not matched_activity or best_score == 0:
        # Nessun match preciso: restituisci tutte le attività con dettagli
        all_lines = []
        for act in activities:
            ap_list = ActivityParticipant.query.filter_by(activity_id=act.id).all()
            all_lines.append(f"\n## {act.name} ({act.date or '?'}) — {len(ap_list)} partecipanti")
            for ap in ap_list:
                all_lines.append(f"  - {ap.last_name} {ap.first_name} ({ap.entity or ''})")
        return 'ALL', '\n'.join(all_lines)

    # Match trovato: dettaglio completo
    ap_list = ActivityParticipant.query.filter_by(activity_id=matched_activity.id).all()
    lines = [f"Attività: {matched_activity.name} ({matched_activity.date or '?'}) — {len(ap_list)} partecipanti"]
    lines.append("Cognome|Nome|Ente|Dieta")
    for ap in ap_list:
        lines.append(f"{ap.last_name or ''}|{ap.first_name or ''}|{ap.entity or ''}|{ap.diet or ''}")
    return matched_activity, '\n'.join(lines)


def _ai_smart_filter(batch_id, question):
    """Filtra i partecipanti dal DB in base alla domanda. Restituisce (rows, filtro_desc, extra_context)."""
    from sqlalchemy import and_
    q_lower = question.lower()
    filters = [RoomingList.import_batch == batch_id]
    desc_parts = []
    extra_context = ''

    # 1. Controlla se è una domanda sulle attività
    act_match, act_data = _ai_detect_activity(question)
    if act_match is not None:
        extra_context = f"\n\nDATI ATTIVITÀ:\n{act_data}"
        # Se è una domanda solo sulle attività, non filtrare il rooming in modo restrittivo
        if act_match != 'ALL':
            desc_parts.append(f"attività={act_match.name}")

    # 2. Rileva hotel nella domanda
    all_hotels = db.session.query(RoomingList.hotel).filter_by(import_batch=batch_id)\
        .distinct().all()
    hotel_names = [h[0] for h in all_hotels if h[0]]
    matched_hotel = None
    for h in hotel_names:
        h_words = h.lower().split()
        for hw in h_words:
            if len(hw) > 3 and hw in q_lower:
                matched_hotel = h
                break
        if matched_hotel:
            break
    if matched_hotel:
        filters.append(RoomingList.hotel == matched_hotel)
        desc_parts.append(f"hotel={matched_hotel}")

    # 3. Rileva notte nella domanda
    matched_night = None
    for field, aliases in AI_NIGHT_MAP.items():
        for alias in aliases:
            if alias.lower() in q_lower:
                matched_night = field
                break
        if matched_night:
            break
    if matched_night:
        filters.append(getattr(RoomingList, matched_night).ilike('Yes'))
        desc_parts.append(f"notte={AI_NIGHT_LABELS[matched_night]}")

    # 4. Rileva CXL / cancellati (attenzione: "attività" contiene "attiv")
    if any(kw in q_lower for kw in ['cxl', 'cancellat', 'annullat']):
        filters.append(RoomingList.registration_state.ilike('CXL'))
        desc_parts.append("stato=CXL")

    # 5. Rileva dieta
    if any(kw in q_lower for kw in ['dieta', 'diete', 'dietetic', 'allergic', 'vegetarian',
                                     'vegan', 'halal', 'kosher', 'intoleran', 'celiac']):
        filters.append(RoomingList.diet_restrictions != None)
        filters.append(RoomingList.diet_restrictions != '')
        desc_parts.append("con dieta")

    # 6. Rileva paese
    if any(kw in q_lower for kw in ['italia', 'italian']):
        filters.append(RoomingList.company_country.ilike('%ital%'))
        desc_parts.append("paese=Italy")

    # 7. Rileva nomi di persone
    import re
    stopwords = {'quale','quali','hotel','albergo','dove','quando','quanti','quante',
                 'chi','cosa','come','sono','nella','nello','nel','della','dello','del',
                 'giorno','notte','data','lista','elenco','partecipanti','persone',
                 'tutti','tutte','con','per','che','non','gli','dei','delle','suo','sua',
                 'signor','signora','sig','può','hanno','anno','questo','questa',
                 'dorme','dormono','sta','stanno','trova','trovano','marzo','aprile',
                 'nomi','nome','cognome','dieta','arrivo','paese','azienda','attività',
                 'attivita','partecipa','iscritto','iscritti','evento','sessione','fai',
                 'workshop','meeting','cena','pranzo','dinner','lunch','gala','tour'}
    for h in hotel_names:
        for hw in h.lower().split():
            if len(hw) > 2:
                stopwords.add(hw)
    words = [w for w in re.findall(r'[a-zàèéìòù]+', q_lower) if len(w) > 2 and w not in stopwords]
    if words and not act_match:
        # Solo se non è una query su attività (evita di filtrare per parole dell'attività)
        from sqlalchemy import or_
        name_conds = []
        for w in words:
            name_conds.append(RoomingList.last_name.ilike(f'%{w}%'))
            name_conds.append(RoomingList.first_name.ilike(f'%{w}%'))
            name_conds.append(RoomingList.company_name.ilike(f'%{w}%'))
        filters.append(or_(*name_conds))
        desc_parts.append(f"cerca={','.join(words)}")

    # Se la domanda è solo sulle attività, non servono tutti i partecipanti rooming
    if act_match and len(filters) == 1:
        # Solo filtro batch_id, non mandare 780 righe
        rows = []
        desc = 'attività'
        return rows, desc, extra_context

    rows = RoomingList.query.filter(and_(*filters))\
        .order_by(RoomingList.last_name, RoomingList.first_name).all()
    desc = ' + '.join(desc_parts) if desc_parts else 'tutti'
    return rows, desc, extra_context


def _ai_format_rows(rows):
    """Formatta righe in CSV compatto."""
    lines = ['Cognome|Nome|Hotel|Stato|Azienda|Paese|CI|CO|Notti|Dieta|Arrivo|Note']
    for r in rows:
        nights = ','.join(lbl for fld, lbl in AI_NIGHT_LABELS.items()
                         if (getattr(r, fld) or '').strip().upper() == 'YES')
        parts = [
            r.last_name or '', r.first_name or '', r.hotel or '',
            r.registration_state or '', r.company_name or '',
            r.company_country or '',
            cell_val('check_in', r), cell_val('check_out', r),
            nights, r.diet_restrictions or '', r.arrival_mode or '',
            r.comment or '',
        ]
        lines.append('|'.join(v.replace('|', '/') for v in parts))
    return '\n'.join(lines)


@rooming_bp.route('/api/ask/<path:batch_id>', methods=['POST'])
def api_ask(batch_id):
    """Risponde a domande in linguaggio naturale sui dati del batch usando Claude."""
    data = request.get_json()
    question = (data.get('question') or '').strip()
    history = data.get('history', [])

    if not question:
        return jsonify({'ok': False, 'error': 'Domanda vuota'})

    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return jsonify({'ok': False, 'error': 'ANTHROPIC_API_KEY non configurata'})

    # Tutti i record per le statistiche
    all_rows = RoomingList.query.filter_by(import_batch=batch_id).all()
    if not all_rows:
        return jsonify({'ok': False, 'error': 'Batch non trovato'})

    summary = _ai_build_summary(all_rows)

    # Filtro intelligente dal DB in base alla domanda
    filtered, filter_desc, extra_context = _ai_smart_filter(batch_id, question)
    filtered_csv = _ai_format_rows(filtered) if filtered else '(nessun partecipante rooming filtrato)'

    system_prompt = f"""Assistente rooming evento. Rispondi in italiano, conciso. Tabelle markdown se utile.
Dati CSV separatore |. Notti=date pernottamento. CXL=cancellato.
I dati mostrati sono GIÀ FILTRATI dal database: questi sono TUTTI i risultati, non un campione.

{summary}

RISULTATI FILTRATI ({filter_desc}): {len(filtered)} partecipanti su {len(all_rows)} totali
{filtered_csv}{extra_context}"""

    messages = []
    for msg in history[-10:]:
        messages.append({'role': msg['role'], 'content': msg['content']})
    messages.append({'role': 'user', 'content': question})

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=2048,
            system=system_prompt,
            messages=messages,
        )
        answer = response.content[0].text
        return jsonify({'ok': True, 'answer': answer})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
